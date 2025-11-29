# -*- coding: utf-8 -*-

"""

Import_absence.py

Lecteur d'absences par couleur depuis un Excel "mois × jours × personnes".



RÈGLES (priorité à la couleur) :

- Jaune  : Absence souhaitée → Journée

- Rouge  : Repos de garde    → Journée

- Vert   : Formation         → Matin (M) / AP (A) / Journée (si vide ou M+A)

- Violet   : Astreinte (soir)  → AP MIDI

- Gris/Autres : ignorés



Onglet = mois (ex. "Août"/"Aout"). A1 = Année (ex. 2025).

B1 = "NOM Prénom". C1..AG1 = "01".."31" (numéros de jour).

Colonne B (à partir de la ligne 2) = une personne par ligne.



Étape 1 : lecture + résumé. L’intégration dans les tableaux viendra à l’étape 2.

"""

from __future__ import annotations



from tkinter import filedialog, messagebox

from datetime import date

from collections import Counter

import unicodedata

import re

import calendar



try:

    from openpyxl import load_workbook

except Exception as e:

    load_workbook = None



# ------------------ Utilitaires de normalisation ------------------ #



def _norm(s):

    """Normalise une chaîne (strip + suppression des accents)."""

    if s is None:

        return ""

    if not isinstance(s, str):

        s = str(s)

    s = s.strip()

    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

    return s



NAME_SPLIT_RE = re.compile(r"[,\n\r\t;/&+]+")



def _split_people(cell_text) -> list[str]:

    if cell_text is None:

        return []

    if not isinstance(cell_text, str):

        cell_text = str(cell_text)

    parts = NAME_SPLIT_RE.split(cell_text)

    cleaned = []

    for part in parts:

        part = part.strip()

        if not part:

            continue

        part = part.strip("\u00A0\u2007\u202F")

        if part:

            cleaned.append(part)

    return cleaned


DAY_LABELS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


MONTHS_FR = {

    "janvier": 1, "fevrier": 2, "février": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,

    "juillet": 7, "aout": 8, "août": 8, "septembre": 9, "octobre": 10, "novembre": 11,

    "decembre": 12, "décembre": 12,

}



def _month_from_sheet(sheet_name: str) -> int | None:

    name = _norm(sheet_name).lower()

    for k, v in MONTHS_FR.items():

        if k in name:

            return v

    m = re.search(r"\b(1[0-2]|[1-9])\b", name)

    if m:

        n = int(m.group(1))

        if 1 <= n <= 12:

            return n

    return None



def _parse_year(ws) -> int | None:

    # Cherche d'abord A1, sinon une année dans la 1re ligne.

    v = ws.cell(1, 1).value

    if isinstance(v, int) and 1900 <= v <= 2100:

        return v

    for c in range(1, 6):

        v = ws.cell(1, c).value

        if isinstance(v, int) and 1900 <= v <= 2100:

            return v

        if isinstance(v, str):

            m = re.search(r"(19|20)\d{2}", v)

            if m:

                return int(m.group(0))

    return None



from datetime import timedelta, date

import tkinter as tk





def _compute_full_weeks(year: int, month: int):

    """

    Calcule les lundis des semaines qui chevauchent le mois (semaines de bord incluses).

    """

    first = date(year, month, 1)

    if month == 12:

        next_month = date(year + 1, 1, 1)

    else:

        next_month = date(year, month + 1, 1)

    last = next_month - timedelta(days=1)



    monday = first - timedelta(days=first.weekday())

    mondays = []

    while monday <= last:

        mondays.append(monday)

        monday += timedelta(days=7)

    return mondays





def _header_day_columns(ws) -> dict[int, int]:

    """Renvoie {col_index: jour_int} pour les colonnes C.. fin où l’en-tête est 01..31."""

    cols = {}

    for c in range(3, ws.max_column + 1):

        v = ws.cell(1, c).value

        if v is None:

            continue

        s = _norm(v)

        m = re.match(r"^0*([1-9]|[12]\d|3[01])$", s)

        if m:

            cols[c] = int(m.group(1))

    return cols



# ------------------ Couleurs → catégories ------------------ #



def _safe_rgb(cell) -> str | None:

    """

    Récupère la couleur de remplissage en hex ARGB 'FFRRGGBB' si disponible.

    Retourne None si pas de couleur exploitable.

    """

    try:

        fill = cell.fill

        ft = getattr(fill, "fill_type", None)

        if not ft or ft == "none":

            return None

        color = getattr(fill, "fgColor", None)

        if color is None:

            return None

        # Cases usuelles : rgb / indexed / theme

        if getattr(color, "type", None) == "rgb" and isinstance(color.rgb, str):

            return color.rgb  # "FFRRGGBB"

        if getattr(color, "type", None) == "indexed":

            # non fiable pour nos besoins

            return None

        if getattr(color, "type", None) == "theme":

            # dépend d'une palette Excel : on ignore

            return None

        if isinstance(getattr(color, "rgb", None), str):

            return color.rgb

    except Exception:

        return None

    return None



def _classify_rgb(rgb_hex: str) -> str | None:

    """

    Classe une couleur Excel en grande famille : YELLOW / RED / GREEN / BROWN / PURPLE / BLUE / GRAY / OTHER.



    Entrées acceptées :

      - "FFRRGGBB" (ARGB)

      - "RRGGBB"   (RGB sans alpha)

      - "#AARRGGBB" ou "#RRGGBB"



    Normalisation :

      - On ajoute FF devant si l'alpha est absent (RRGGBB).

      - Si l'alpha n'est pas FF, on lit quand même les composantes R,G,B.

    """

    if not isinstance(rgb_hex, str) or not rgb_hex.strip():

        return None



    s = rgb_hex.strip().lstrip("#")

    if len(s) == 6:          # RRGGBB

        s = "FF" + s

    elif len(s) != 8:        # format inattendu

        return None



    s = s.upper()

    try:

        r = int(s[2:4], 16)

        g = int(s[4:6], 16)

        b = int(s[6:8], 16)

    except Exception:

        return None



    # --- Tables de couleurs "exactes" rencontrées dans Excel/Office ---

    KNOWN_BROWNS = {

        "FFA52A2A",  # Brown

        "FF8B4513",  # SaddleBrown

        "FF7F6000",  # Office orange-ish

        "FF804000",  # Office brown-ish

        "FFC55A11",  # Accent 6 foncé

        "FF964B00",  # Burnt Orange

        "FF7E3F00",  # Variante foncée fréquente

    }

    if s in KNOWN_BROWNS:

        return "BROWN"



    KNOWN_PURPLES = {

        "FF800080",  # Purple standard (#800080)

        "FF7030A0",  # Office Purple (souvent "Accent 5 - Darker 25%")

        "FF9933FF",  # Violet soutenu (variante fréquente)

        "FF6600CC",  # Violet/Indigo (variante vive)

    }

    if s in KNOWN_PURPLES:

        return "PURPLE"



    # --- Heuristiques robustes par composantes ---

    # Gris : canaux proches

    if max(abs(r - g), abs(r - b), abs(g - b)) < 18:

        return "GRAY"



    # Jaune vif

    if r > 200 and g > 200 and b < 140:

        return "YELLOW"



    # Marron : rouge dominant, vert moyen, bleu bas

    if (r > 110 and b < 100 and 40 <= g <= 190 and

            r >= g + 20 and r >= b + 40 and g < 210):

        return "BROWN"



    # Violet/Magenta : rouge et bleu hauts, vert bas

    if r >= 120 and b >= 120 and g <= 90:

        return "PURPLE"



    # Rouge franc

    if r > 200 and g < 110 and b < 110:

        return "RED"



    # Vert franc

    if g > 130 and g > r + 15 and g > b + 15:

        return "GREEN"



    # Bleu franc

    if b > 130 and b > r + 15 and b > g + 15:

        return "BLUE"



    return "OTHER"



def _get_cell_category(cell) -> str | None:

    """

    Renvoie la catégorie couleur brute:

    'YELLOW' / 'RED' / 'GREEN' / 'BROWN' / 'BLUE' / 'GRAY' / 'OTHER' / None



    ⚠️ Stratégie (plus stricte) :

      - On ignore désormais les motifs Excel (ex. 'gray125') -> évite les faux "jours fermés".

      - On NE déduit PLUS 'GRAY' depuis les couleurs de thème Office (theme/tint).

      - On ne considère 'GRAY' que si:

          * la cellule a un REMPLISSAGE SOLIDE en RGB qui est effectivement gris (canaux proches), ou

          * certaines couleurs 'indexed' historiquement grises.

      - On conserve en revanche les détections utiles pour les absences individuelles:

          * 'GREEN' depuis certains thèmes verts (formation),

          * 'BROWN' (astreinte) et 'PURPLE' pour distinguer de GREEN.

    """

    f = getattr(cell, "fill", None)

    if not f:

        return None



    # 1) N'interpréter que les remplissages SOLIDES.

    #    -> on ignore les motifs (ex. 'gray125') qui causent des faux positifs.

    ft = getattr(f, "fill_type", None)

    if not isinstance(ft, str) or ft.lower() in (None, "none"):

        return None

    if ft.lower() != "solid":

        return None



    # 2) Inspecter la couleur en priorité via RGB ; fallback sur indexed / theme (sans GRAY via theme)

    for attr in ("start_color", "fgColor", "end_color", "bgColor"):

        c = getattr(f, attr, None)

        if c is None:

            continue

        ctype = getattr(c, "type", None)



        # RGB direct -> classification robuste (gris si R,G,B proches, etc.)

        if ctype == "rgb":

            rgb = getattr(c, "rgb", None)

            cat = _classify_rgb(rgb)  # YELLOW/RED/GREEN/BROWN/PURPLE/BLUE/GRAY/OTHER

            if cat:

                return cat



        # Couleurs indexées -> certains indices sont réellement gris

        if ctype == "indexed":

            idx = getattr(c, "indexed", None)

            if idx in (22, 23, 24, 25, 26, 27, 28, 64):

                return "GRAY"



        # Thèmes Office -> NE PAS en déduire GRAY ; on garde uniquement les cas utiles

        if ctype == "theme":

            theme = getattr(c, "theme", None)

            # garder GREEN pour formation (accents verts courants)

            if theme in (6, 7):

                return "GREEN"

            # garder PURPLE pour le distinguer de GREEN

            if theme == 8:

                return "PURPLE"

            # garder BROWN (astreinte)

            if theme == 9:

                return "BROWN"



    return None





def _classify_cell(cell) -> tuple[str, str] | None:

    """

    Retourne (reason, span) ou None.

    - reason ∈ {"souhait","repos","formation","astreinte"}

    - span   ∈ {"MATIN","AP MIDI","Journée"}



    Règle prioritaire = la couleur.

    VERT (formation) — logique ultra simple :

      - si la cellule (après trim) est exactement 'm'/'M' → MATIN

      - si exactement 'a'/'A' → AP MIDI

      - sinon (vide ou autre) → Journée

    """



    def _span_from_green_text(raw) -> str:

        # Normalisation minimale : string, suppression espaces invisibles fréquents

        if raw is None:

            return "Journée"

        s = str(raw)

        # retirer espaces insécables/zero-width

        s = s.replace("\u00A0", " ").replace("\u200B", "")

        s = s.strip()



        if len(s) == 1:

            ch = s.lower()

            if ch == "m":

                return "MATIN"

            if ch == "a":

                return "AP MIDI"

        # tout le reste (vide, 'ma', 'formation', etc.) -> Journée

        return "Journée"



    cat = _get_cell_category(cell)

    if not cat:

        return None



    if cat == "YELLOW":

        return ("souhait", "Journée")

    if cat == "RED":

        return ("repos", "Journée")

    if cat == "GREEN":

        return ("formation", _span_from_green_text(cell.value))

    if cat in ("BROWN", "PURPLE"):

        # Astreinte du soir → indisponible l'après-midi

        return ("astreinte", "AP MIDI")



    # GRAY/BLUE/OTHER → pas une absence individuelle

    return None





# ------------------ Helpers matching personnes → initiales + injection ------------------ #



def _norm_init(s: str) -> str:

    """Normalise une initiale : supprime accents/espaces/points, uppercase."""

    s = _norm(s).upper()

    s = s.replace(".", "").replace(" ", "").replace("-", "")

    return s



def _guess_initials_from_name(person: str) -> list[str]:

    """

    Devine des initiales possibles à partir de 'NOM Prénom' (ou 'Prénom NOM').

    Retourne une liste (ordre de préférence).

    """

    if not person:

        return []

    txt = _norm(person).strip()

    if not txt:

        return []

    parts = [p for p in re.split(r"[ \t]+", txt) if p]

    # Gestion des tirets (Jean-Pierre => JP)

    def first_letter_token(tok: str) -> str:

        subs = [t for t in tok.split("-") if t]

        return "".join(sub[0] for sub in subs if sub)



    # Heuristique principale : Excel annoncé comme "NOM Prénom"

    if len(parts) >= 2:

        nom = parts[0]

        prenom = parts[-1]

        i_nom = first_letter_token(nom)

        i_pre = first_letter_token(prenom)

        cands = [

            (i_nom[:1] + i_pre[:1]).upper(),   # D + A -> DA

            (i_pre[:1] + i_nom[:1]).upper(),   # A + D -> AD

        ]

        # Variantes robustes si noms composés

        if len(i_nom) > 1:

            cands.append((i_nom[:2] + i_pre[:1]).upper())

        if len(i_pre) > 1:

            cands.append((i_nom[:1] + i_pre[:2]).upper())

        # Uniqueness

        out, seen = [], set()

        for c in cands:

            n = _norm_init(c)

            if n and n not in seen:

                out.append(c)

                seen.add(n)

        return out

    # Un seul token → on tente deux lettres

    tok = parts[0]

    if len(tok) >= 2:

        return [tok[0].upper() + tok[1].upper()]

    return [tok[0].upper()]



def _find_row_by_initials(constraints_app, initials: str):

    """

    Retourne (row_index, row_widgets) dont la col 0 == initials (insensible aux accents/casse),

    ou (None, None) si absent.

    """

    target = _norm_init(initials)

    for idx, row in enumerate(getattr(constraints_app, "rows", [])):

        try:

            cur = row[0].get().strip()

        except Exception:

            cur = ""

        if _norm_init(cur) == target:

            return idx, row

    return None, None



def _ensure_row_exists_in_all_weeks(tabs_data, initials: str):

    """

    S'assure que la personne 'initials' existe dans le tableau de contraintes de CHAQUE semaine.

    Si manquante, crée une ligne et positionne la colonne 0.

    Retourne la liste des index de ligne (par semaine).

    """

    row_idxs = []

    for (_g, c, _s) in tabs_data:

        idx, row = _find_row_by_initials(c, initials)

        if idx is None:

            # créer une ligne à la fin

            c.add_row()

            row = c.rows[-1]

            try:

                row[0].delete(0, "end")

                row[0].insert(0, initials)

            except Exception:

                pass

            idx = len(c.rows) - 1

        row_idxs.append(idx)

    return row_idxs



def _merge_span(current: str, new: str) -> str:

    """

    Fusionne deux spans ("", MATIN, AP MIDI, Journée) → Journée si M + A.

    Ne 'rétrécit' jamais : Journée reste Journée.

    """

    def norm(s):

        s = (s or "").strip().upper()

        s = s.replace("JOURNÉE", "JOURNEE")

        return s

    cur = norm(current)

    nxt = norm(new)

    if not nxt:

        return current  # rien à ajouter

    if cur == "JOURNEE" or nxt == "JOURNEE":

        return "Journée"

    if (cur == "MATIN" and nxt == "AP MIDI") or (cur == "AP MIDI" and nxt == "MATIN"):

        return "Journée"

    if not cur:

        return "MATIN" if nxt == "MATIN" else ("AP MIDI" if nxt == "AP MIDI" else "Journée")

    # même type, on garde

    return "MATIN" if cur == "MATIN" else ("AP MIDI" if cur == "AP MIDI" else "Journée")





# ------------------ Parsing d'une feuille/mois ------------------ #



def _parse_month_sheet(ws) -> dict:

    """

    ws -> dict {

        'year', 'month', 'sheet',

        'people_all': [name, ...],

        'people_all_count': int,

        'people': { name: [ {date, reason, span}, ... ] },

        'people_with_marks_count': int,

        'entries': [...],

        'summary': { reason: count, ... },

        'closed_dates': [iso-date, ...]   # jours de semaine détectés gris dans l'Excel

    }

    """

    year = _parse_year(ws)

    month = _month_from_sheet(ws.title)

    if year is None or month is None:

        raise ValueError(f"Mois/année introuvables (Feuille '{ws.title}'). "

                         f"Vérifie A1 (année) et le nom de l'onglet (mois).")



    day_cols = _header_day_columns(ws)

    sorted_cols = sorted(day_cols.items())



    cal = calendar.Calendar(firstweekday=0)

    flat_dates = []

    for week in cal.monthdatescalendar(year, month):

        flat_dates.extend(week)



    needed = len(sorted_cols)

    col_dates: dict[int, date] = {}

    week_mondays: list[date] = []

    if needed:

        target_len = needed + 14  # marges supplémentaires pour trouver l'alignement

        while len(flat_dates) < target_len:

            last = flat_dates[-1]

            next_week_start = last + timedelta(days=1)

            flat_dates.extend([next_week_start + timedelta(days=i) for i in range(7)])

        day_values = [day for _col, day in sorted_cols]

        flat_days = [d.day for d in flat_dates]

        start_idx = None

        for idx in range(0, len(flat_days) - needed + 1):

            if flat_days[idx:idx + needed] == day_values:

                start_idx = idx

                break

        if start_idx is not None:

            for offset, (col_idx, _day) in enumerate(sorted_cols):

                col_dates[col_idx] = flat_dates[start_idx + offset]

        else:

            first_one = next((idx for idx, (_, day) in enumerate(sorted_cols) if day == 1), None)

            prev_year = year - 1 if month == 1 else year

            prev_month = 12 if month == 1 else month - 1

            current_year = year

            current_month = month

            prev_day = None

            for idx, (col_idx, day) in enumerate(sorted_cols):

                if first_one is not None and idx < first_one and day > 20:

                    try:

                        col_dates[col_idx] = date(prev_year, prev_month, day)

                    except ValueError:

                        continue

                    prev_day = day

                    continue

                if prev_day is not None and day < prev_day:

                    if current_month == 12:

                        current_year += 1

                        current_month = 1

                    else:

                        current_month += 1

                try:

                    col_dates[col_idx] = date(current_year, current_month, day)

                except ValueError:

                    continue

                prev_day = day



        if col_dates:

            seen_mondays = set()

            for actual_date in sorted(col_dates.values()):

                monday = actual_date - timedelta(days=actual_date.weekday())

                if monday not in seen_mondays:

                    seen_mondays.add(monday)

                    week_mondays.append(monday)



    people = {}

    people_all = []

    entries = []

    counts = Counter()

    gray_counts = Counter()



    for r in range(2, ws.max_row + 1):

        person = ws.cell(r, 2).value

        if not isinstance(person, str) or not person.strip():

            continue

        person = person.strip()

        people_all.append(person)



        per_list = []

        for col_idx, _day in sorted_cols:

            actual_date = col_dates.get(col_idx)

            if actual_date is None:

                continue

            cell = ws.cell(r, col_idx)



            cat = _get_cell_category(cell)

            if cat == "GRAY":

                gray_counts[actual_date] += 1

                continue



            res = _classify_cell(cell)

            if not res:

                continue

            reason, span = res

            entry = {"person": person, "date": actual_date.isoformat(), "reason": reason, "span": span}

            per_list.append(entry)

            entries.append(entry)

            counts[reason] += 1



        if per_list:

            people[person] = per_list



    closed_dates = []

    total_people = len(people_all) if people_all else 1

    for day_date, gcount in gray_counts.items():

        if day_date.weekday() < 5 and gcount / total_people >= 0.75:

            closed_dates.append(day_date.isoformat())



    return {

        "year": year,

        "month": month,

        "sheet": ws.title,

        "people_all": people_all,

        "people_all_count": len(people_all),

        "people": people,

        "people_with_marks_count": len(people),

        "entries": entries,

        "summary": dict(counts),

        "closed_dates": closed_dates,

        "week_mondays": week_mondays,

    }



def parse_absence_workbook(xlsx_path: str) -> list[dict]:

    """

    Lit le classeur et renvoie une liste de mois parsés (un dict par onglet/mois).

    """

    if load_workbook is None:

        raise ImportError(

            "openpyxl est requis pour lire les couleurs Excel. "

            "Installe-le (pip install openpyxl) puis réessaie."

        )

    wb = load_workbook(xlsx_path, data_only=True)

    months = []

    for name in wb.sheetnames:

        ws = wb[name]

        # On ne garde que les onglets qui ressemblent à un mois

        if _month_from_sheet(name) is None:

            continue

        months.append(_parse_month_sheet(ws))

    if not months:

        raise ValueError("Aucun onglet de mois valide trouvé (ex. 'Janvier', 'Août', ...).")

    return months



# ------------------ Crochets UI (étape 1 = lecture + résumé) ------------------ #



def import_absences_from_excel(root, notebook, tabs_data):

    """

    Import d’un fichier d’absences Excel avec mapping robuste (aucune création de lignes) :

    - On matche UNIQUEMENT les noms déjà présents en colonne 0 du tableau de contraintes

      avec les noms de l’Excel.

    - Les candidats Excel viennent de people_all (TOUS les noms), pas seulement des personnes

      avec une marque (people). Cela résout les cas comme "IMPERADORI L." non proposés.

    - Matching multi-niveaux : normalisation stricte, ensemble de tokens (stopwords ignorés),

      initiales étendues, et score composite (séquence + tokens).

    - En cas d’ambiguïté / échec : boîte de dialogue avec shortlist + % et option "Ignorer".

    - Injection des absences (M/AP/Journée) sans modifier le nombre de lignes.

    """

    from tkinter import filedialog, messagebox, ttk

    from collections import Counter, defaultdict

    import tkinter as tk

    from datetime import date, timedelta

    import difflib

    import re



    STOPWORDS = {

        "DE","DU","DES","D","LE","LA","LES","DEL","DELA","DA","DOS","DO",

        "VAN","VON","DI","DELLA","MC","MAC"

    }

    INVISIBLES = {"\u00A0","\u2007","\u202F","\u200B","\u200C","\u200D","\uFEFF"}



    def _get_toggle_state(toggle) -> str:

        try:

            value = toggle._var.get()

        except Exception:

            try:

                value = toggle.cget("text")

            except Exception:

                value = ""

        return (value or "").strip()



    def _apply_toggle_state(toggle, state: str):

        state = state or ""

        try:

            setter = getattr(toggle, "set_state", None)

            if callable(setter):

                setter(state)

                return

        except Exception:

            pass

        try:

            toggle._var.set(state)

            toggle.config(text=state)

        except Exception:

            pass



    def _capture_week_absences(constraints_app) -> list[list[str]]:

        baselines = []

        for row in getattr(constraints_app, "rows", []):

            day_states = []

            for day_idx in range(7):

                try:

                    cell_tuple = row[4 + day_idx]

                except Exception:

                    cell_tuple = None

                if isinstance(cell_tuple, tuple) and cell_tuple:

                    toggle = cell_tuple[0]

                    state = _get_toggle_state(toggle)

                else:

                    state = ""

                day_states.append(state)

            baselines.append(day_states)

        return baselines



    def _restore_week_absences(constraints_app, baseline: list[list[str]]):

        rows = getattr(constraints_app, "rows", [])

        for row_idx, day_states in enumerate(baseline):

            if row_idx >= len(rows):

                break

            row = rows[row_idx]

            for day_idx, state in enumerate(day_states):

                try:

                    cell_tuple = row[4 + day_idx]

                except Exception:

                    continue

                if not (isinstance(cell_tuple, tuple) and cell_tuple):

                    continue

                toggle = cell_tuple[0]

                _apply_toggle_state(toggle, state)



    def _strip_invisibles(s: str) -> str:

        if not isinstance(s, str): return ""

        for ch in INVISIBLES:

            s = s.replace(ch, " ")

        return s



    def _norm_name(s: str) -> str:

        s = _strip_invisibles(s)

        s = _norm(s).upper()

        s = s.replace("’", "'").replace("‐","-").replace("–","-").replace("—","-")

        s = re.sub(r"[.\-_/(){}\[\],;:!?\t]+", " ", s)

        s = re.sub(r"\s+", " ", s).strip()

        return s



    def _tokens(s: str) -> list[str]:

        txt = _norm_name(s)

        if not txt: return []

        toks = []

        for t in txt.split(" "):

            if not t: continue

            parts = [p for p in t.split("-") if p]

            for p in parts:

                if p and p not in STOPWORDS:

                    toks.append(p)

        return toks



    def _tokens_len_ge2(s: str) -> set[str]:

        return {t for t in _tokens(s) if len(t) >= 2}



    def _token_key(s: str) -> str:

        toks = sorted(set(_tokens(s)))

        return " ".join(toks)



    def _token_set_ratio(a: str, b: str) -> float:

        A, B = set(_tokens(a)), set(_tokens(b))

        if not A or not B: return 0.0

        inter = len(A & B)

        return (2.0 * inter) / (len(A) + len(B))



    def _seq_ratio(a: str, b: str) -> float:

        return difflib.SequenceMatcher(None, _norm_name(a), _norm_name(b)).ratio()



    def _combined_score(a: str, b: str) -> float:

        return 0.60 * _seq_ratio(a, b) + 0.40 * _token_set_ratio(a, b)



    INITIALS_AUTO_SCORE = 0.60

    INITIALS_SHORTLIST_SCORE = 0.50

    INITIALS_SHORTLIST_LIMIT = 6



    def _long_tokens(name: str) -> list[str]:

        return [t for t in _tokens(name) if len(t) >= 3]



    def _share_token_len_ge3(a: str, b: str) -> bool:

        return bool(set(_long_tokens(a)) & set(_long_tokens(b)))



    def _has_conflicting_long_tokens(a: str, b: str) -> bool:

        tokens_a = _long_tokens(a)

        tokens_b = _long_tokens(b)

        if not tokens_a or not tokens_b:

            return False

        def _matches(tok: str, pool: list[str]) -> bool:

            prefix = tok[:3]

            for other in pool:

                if tok == other or (prefix and prefix == other[:3]):

                    return True

            return False

        extra_a = [tok for tok in tokens_a if not _matches(tok, tokens_b)]

        extra_b = [tok for tok in tokens_b if not _matches(tok, tokens_a)]

        return bool(extra_a and extra_b)



    def _initial_variants(full: str) -> set[str]:

        txt = _norm_name(full)

        if not txt: return set()

        toks = txt.split()

        def first_letters(token: str) -> str:

            parts = [p for p in token.split("-") if p]

            return "".join(p[0] for p in parts if p)

        out = set()

        if len(toks) == 1:

            t = toks[0]

            if len(t) >= 1: out.add(t[:1])

            if len(t) >= 2: out.add(t[:2])

            return {x.replace(" ","").upper() for x in out}

        first = first_letters(toks[0]); last = first_letters(toks[-1])

        if first and last:

            out.add((first[:1] + last[:1]))

            out.add((last[:1] + first[:1]))

            if len(first) >= 2: out.add(first[:2] + last[:1])

            if len(last)  >= 2: out.add(first[:1] + last[:2])

        out.add((toks[0] + " " + (last[:1] if last else "")).replace(" ",""))

        if first: out.add((first[:1] + " " + toks[-1]).replace(" ",""))

        if first or last: out.add((first or "") + (last or ""))

        return {x.upper() for x in out if x}



    def _merge_span(current: str, new: str) -> str:

        def norm(s):

            return (s or "").strip().upper().replace("JOURNÉE","JOURNEE")

        cur = norm(current); nxt = norm(new)

        if not nxt: return current

        if cur == "JOURNEE" or nxt == "JOURNEE": return "Journée"

        if (cur == "MATIN" and nxt == "AP MIDI") or (cur == "AP MIDI" and nxt == "MATIN"):

            return "Journée"

        if not cur:

            return "MATIN" if nxt == "MATIN" else ("AP MIDI" if nxt == "AP MIDI" else "Journée")

        return "MATIN" if cur == "MATIN" else ("AP MIDI" if cur == "AP MIDI" else "Journée")



    def _find_row_by_display_name(constraints_app, display: str):

        target = _norm_name(display)

        for idx, row in enumerate(getattr(constraints_app, "rows", [])):

            try:

                cur = row[0].get().strip()

            except Exception:

                cur = ""

            if _norm_name(cur) == target:

                return idx, row

        return None, None



    # ---------- Sélection du fichier & parsing ----------

    path = filedialog.askopenfilename(

        title="Sélectionner le fichier d'absences (Excel)",

        filetypes=[("Fichiers Excel", "*.xlsx *.xlsm *.xltx *.xltm"), ("Tous les fichiers", "*.*")]

    )

    if not path: return

    try:

        months = parse_absence_workbook(path)

    except Exception as e:

        messagebox.showerror("Import Absences", f"Échec de lecture du fichier :\n{e}")

        return

    if not months:

        messagebox.showerror("Import Absences", "Aucun onglet de mois valide trouvé.")

        return



    def _mois_fr(n):

        mapping = {1:"Janvier",2:"Février",3:"Mars",4:"Avril",5:"Mai",6:"Juin",

                   7:"Juillet",8:"Août",9:"Septembre",10:"Octobre",11:"Novembre",12:"Décembre"}

        return mapping.get(n, f"Mois {n}")



    def _ask_user_to_pick_month(parent, months_list):

        if not months_list: return None

        if len(months_list) == 1: return months_list[0]

        win = tk.Toplevel(parent); win.title("Choisir le mois à importer")

        try: win.geometry(f"+{parent.winfo_pointerx()}+{parent.winfo_pointery()}")

        except Exception: pass

        tk.Label(win, text="Sélectionnez un onglet (mois) :", font=("Arial", 10, "bold")).pack(padx=10, pady=(10,5))

        lst = tk.Listbox(win, height=min(12, len(months_list)), width=42); lst.pack(fill="both", expand=True, padx=10, pady=5)

        for m_ in months_list:

            lst.insert(tk.END, f"{m_['sheet']} — {_mois_fr(m_['month'])} {m_['year']}")

        chosen = {"idx": None}

        def on_ok(_evt=None):

            sel = lst.curselection()

            if not sel:

                messagebox.showinfo("Choix requis", "Veuillez sélectionner un mois.")

                return

            chosen["idx"] = int(sel[0]); win.destroy()

        def on_cancel(): chosen["idx"] = None; win.destroy()

        btnf = tk.Frame(win); btnf.pack(pady=(5,10))

        tk.Button(btnf, text="OK", command=on_ok, width=10).pack(side="left", padx=5)

        tk.Button(btnf, text="Annuler", command=on_cancel, width=10).pack(side="left", padx=5)

        lst.bind("<Double-Button-1>", on_ok)

        win.transient(parent); win.grab_set(); parent.wait_window(win)

        if chosen["idx"] is None: return None

        return months_list[chosen["idx"]]



    m = _ask_user_to_pick_month(root, months)

    if m is None: return



    # ---------- Semaines + ajustement onglets ----------

    weeks_from_excel = [w for w in (m.get("week_mondays") or []) if isinstance(w, date)]

    # Unicity + ordre préservé (Excel → colonnes → semaines)

    weeks = list(dict.fromkeys(weeks_from_excel))

    if not weeks:

        weeks = _compute_full_weeks(m["year"], m["month"])

    if not weeks:

        messagebox.showerror("Import Absences", "Aucune semaine complète dans ce mois.")

        return



    import_start = weeks[0]

    import_end = weeks[-1] + timedelta(days=6)



    target = len(weeks); existing = len(tabs_data)

    try:

        from Full_GUI import add_new_week, create_single_week

    except Exception:

        add_new_week = None; create_single_week = None



    if existing == 0:

        if create_single_week is None:

            messagebox.showerror("Import Absences", "Impossible de créer un onglet (create_single_week indisponible).")

            return

        frame = tk.Frame(notebook); frame.pack(fill="both", expand=True)

        g, c, s = create_single_week(frame)

        tabs_data.append((g, c, s))

        notebook.add(frame, text="Semaine 1"); existing = 1



    if existing < target:

        for _ in range(target - existing):

            if add_new_week is not None: add_new_week()

            elif create_single_week is not None:

                frame = tk.Frame(notebook); frame.pack(fill="both", expand=True)

                g, c, s = create_single_week(frame)

                tabs_data.append((g, c, s))

                notebook.add(frame, text=f"Semaine {len(tabs_data)}")

            else:

                messagebox.showerror("Import Absences", "Aucune méthode pour créer des semaines.")

                return

            

    elif existing > target:

        # Détruire VRAIMENT les derniers onglets en trop (et leur contenu)

        for _ in range(existing - target):

            idx = len(tabs_data) - 1

            try:

                # Récupère l'identifiant tk du tab puis son widget

                tab_id = notebook.tabs()[idx]

                w = notebook.nametowidget(tab_id)  # équiv. root.nametowidget(...)

                # Retire du notebook et détruit le contenu pour libérer la mémoire

                notebook.forget(tab_id)

                try:

                    w.destroy()

                except Exception:

                    pass

            except Exception:

                pass

            # Retire les références Python

            try:

                tabs_data.pop(idx)

            except Exception:

                pass





    # Renommer + fermeture jours gris

    week_absence_baselines = [

        _capture_week_absences(constraints_app)

        for (_g, constraints_app, _s) in tabs_data

    ]



    closed = {date.fromisoformat(d) for d in m.get("closed_dates", [])}

    closed_labels = []

    for i, monday in enumerate(weeks):

        g, _c, s = tabs_data[i]

        g.week_label.config(text=f"Semaine du {monday.strftime('%d/%m/%Y')}")

        for j in range(7):

            d = monday + timedelta(days=j)

            if d < import_start or d > import_end:

                continue

            if d in closed:

                closed_labels.append(d.strftime("%d/%m"))

                total_rows = len(getattr(g, "table_entries", []))

                for r in range(total_rows):

                    try:

                        entry = g.table_entries[r][j]

                        entry.config(state="normal"); entry.delete(0, "end")

                    except Exception: pass

                    try:

                        g.cell_availability[(r, j)] = False; g.update_cell(r, j)

                    except Exception: pass

        try: s.update_counts()

        except Exception: pass

        try: g.update_colors(None)

        except Exception: pass

        if i < len(week_absence_baselines):

            _restore_week_absences(_c, week_absence_baselines[i])

        try: g.auto_resize_all_columns()

        except Exception: pass



    # ---------- Référentiels Contraintes & Excel ----------

    # Contraintes (noms uniques sur toutes semaines)

    constraint_display_names = []

    seen_norm = set()

    for (_g, c, _s) in tabs_data:

        for row in getattr(c, "rows", []):

            try:

                disp = row[0].get().strip()

            except Exception:

                disp = ""

            if not disp: continue

            nd = _norm_name(disp)

            if nd not in seen_norm:

                seen_norm.add(nd); constraint_display_names.append(disp)



    # Excel : distinguer personnes avec marques vs tous les noms

    excel_people_with_marks = m.get("people", {}) or {}

    excel_all_names = list(m.get("people_all", [])) or list(excel_people_with_marks.keys())



    # Index sur TOUS les noms Excel (people_all) pour le matching/candidats

    from collections import defaultdict

    excel_norm_to_names   = defaultdict(list)

    excel_tokenkey_to     = defaultdict(list)

    excel_initials_to     = defaultdict(set)

    excel_token_to_names  = defaultdict(set)



    for name in excel_all_names:

        nn = _norm_name(name)

        excel_norm_to_names[nn].append(name)

        tk_key = _token_key(name)

        if tk_key:

            excel_tokenkey_to[tk_key].append(name)

        for iv in _initial_variants(name):

            excel_initials_to[iv].add(name)

        for t in _tokens_len_ge2(name):

            excel_token_to_names[t].add(name)



    # ---------- Matching Contraintes → Excel (sans création) ----------

    auto_map, to_confirm, to_prompt = {}, [], []



    for disp in constraint_display_names:

        nd  = _norm_name(disp)

        tkd = _token_key(disp)

        ivd = _initial_variants(disp)



        # (1) Égalité normalisée stricte (couvre "IMPERADORI L." vs "Imperadori L")

        if nd in excel_norm_to_names:

            cands = list(dict.fromkeys(excel_norm_to_names[nd]))

            if len(cands) == 1:

                auto_map[disp] = cands[0]; continue

            else:

                to_confirm.append((disp, cands)); continue



        # (2) Ensemble de tokens (ordre libre, stopwords ignorés)

        if tkd in excel_tokenkey_to:

            cands = list(dict.fromkeys(excel_tokenkey_to[tkd]))

            if len(cands) == 1:

                auto_map[disp] = cands[0]; continue

            else:

                to_confirm.append((disp, cands)); continue



        # (3) Initiales

        hits = set()

        for token in ivd:

            if token in excel_initials_to:

                hits |= excel_initials_to[token]

        if hits:

            ordered_hits = sorted(hits, key=_norm_name)

            scored_hits = []

            for nm in ordered_hits:

                if not _share_token_len_ge3(disp, nm):

                    continue

                if _has_conflicting_long_tokens(disp, nm):

                    continue

                score = _combined_score(disp, nm)

                scored_hits.append((nm, score))

            if len(scored_hits) == 1 and scored_hits[0][1] >= INITIALS_AUTO_SCORE:

                auto_map[disp] = scored_hits[0][0]; continue

            if scored_hits:

                scored_hits.sort(key=lambda x: (-x[1], _norm_name(x[0])))

                shortlist_fmt = []

                for idx, (nm, score) in enumerate(scored_hits):

                    if idx >= INITIALS_SHORTLIST_LIMIT:

                        break

                    if idx > 0 and score < INITIALS_SHORTLIST_SCORE:

                        break

                    shortlist_fmt.append(f"{nm} ({int(score*100)}%)")

                if shortlist_fmt:

                    to_confirm.append((disp, shortlist_fmt)); continue



        # (4) Fuzzy : shortlist scorée sur un sous-ensemble par tokens (>=2), sinon tous les noms

        cand_names = set()

        for t in _tokens_len_ge2(disp):

            cand_names |= excel_token_to_names.get(t, set())

        if not cand_names:

            cand_names = set(excel_all_names)



        scored = [(nm, _combined_score(disp, nm)) for nm in cand_names]

        scored.sort(key=lambda x: (-x[1], _norm_name(x[0])))



        if scored:

            top_name, top_sc = scored[0]

            second_sc = scored[1][1] if len(scored) >= 2 else 0.0

            if top_sc >= 0.92 and (top_sc - second_sc) >= 0.05:

                auto_map[disp] = top_name

            else:

                shortlist = [(nm, sc) for (nm, sc) in scored if sc >= 0.82][:7]

                if shortlist:

                    to_confirm.append((disp, [f"{nm} ({int(sc*100)}%)" for nm, sc in shortlist]))

                else:

                    to_prompt.append(disp)

        else:

            to_prompt.append(disp)



    # ---------- Boîte de dialogue ----------

    def _resolve_with_dialog(parent, items_with_scores, items_no_candidates, excel_all_names):

        if not items_with_scores and not items_no_candidates:

            return {}



        win = tk.Toplevel(parent); win.title("Valider les correspondances (Contraintes → Excel)")

        try: win.geometry(f"+{parent.winfo_rootx()+40}+{parent.winfo_rooty()+40}")

        except Exception: pass



        canvas = tk.Canvas(win, borderwidth=0, highlightthickness=0)

        vbar   = tk.Scrollbar(win, orient="vertical", command=canvas.yview)

        frame  = tk.Frame(canvas)

        canvas.create_window((0,0), window=frame, anchor="nw")

        canvas.configure(yscrollcommand=vbar.set)

        def _on_cfg(_e=None): canvas.configure(scrollregion=canvas.bbox("all"))

        frame.bind("<Configure>", _on_cfg)

        canvas.grid(row=0, column=0, sticky="nsew"); vbar.grid(row=0, column=1, sticky="ns")

        win.grid_rowconfigure(0, weight=1); win.grid_columnconfigure(0, weight=1)



        tk.Label(frame, text="Nom dans Contraintes", font=("Arial", 9, "bold")).grid(row=0, column=0, padx=6, pady=4, sticky="w")

        tk.Label(frame, text="Nom Excel correspondant", font=("Arial", 9, "bold")).grid(row=0, column=1, padx=6, pady=4, sticky="w")



        # Marquer visuellement ceux qui ont des absences (★)

        excel_names_marked = []

        marks_set = set(excel_people_with_marks.keys())

        for nm in sorted(excel_all_names, key=_norm_name):

            label = nm + (" ★" if nm in marks_set else "")

            excel_names_marked.append(label)



        rows_vars = []

        row_idx = 1

        for disp, labeled in items_with_scores:

            tk.Label(frame, text=disp).grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")

            values = ["⟶ Ignorer"] + list(labeled)

            var = tk.StringVar(value=values[0])

            cb  = ttk.Combobox(frame, values=values, textvariable=var, width=50, state="readonly")

            cb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="w")

            rows_vars.append((disp, var, "scored"))

            row_idx += 1



        for disp in items_no_candidates:

            tk.Label(frame, text=disp).grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")

            values = ["⟶ Ignorer"] + excel_names_marked

            var = tk.StringVar(value=values[0])

            cb  = ttk.Combobox(frame, values=values, textvariable=var, width=50, state="readonly")

            cb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="w")

            rows_vars.append((disp, var, "all"))

            row_idx += 1



        chosen = {}

        def on_ok():

            for disp, var, kind in rows_vars:

                val = var.get()

                if val == "⟶ Ignorer":

                    chosen[disp] = None

                else:

                    # nettoyer: retirer " (xx%)" et le marqueur " ★"

                    base = re.sub(r"\s*\(\d+%\)$", "", val).rstrip()

                    if base.endswith(" ★"): base = base[:-2].rstrip()

                    chosen[disp] = base

            win.destroy()



        btnf = tk.Frame(win); btnf.grid(row=1, column=0, columnspan=2, pady=6)

        tk.Button(btnf, text="OK", command=on_ok, width=12).pack()



        win.transient(parent); win.grab_set(); parent.wait_window(win)

        return chosen



    manual_choices = _resolve_with_dialog(root, to_confirm, to_prompt, excel_all_names)



    # Final map : chaque nom Contraintes -> nom Excel ou None

    final_map = {}

    for disp in constraint_display_names:

        if disp in auto_map: final_map[disp] = auto_map[disp]

        elif disp in manual_choices: final_map[disp] = manual_choices[disp]

        else: final_map[disp] = None



    # ---------- Injection des absences (uniquement celles réellement présentes) ----------

    for i, monday in enumerate(weeks):

        g, c, s = tabs_data[i]

        week_start = monday

        week_end   = monday + timedelta(days=6)



        for row in getattr(c, "rows", []):

            try:

                disp = row[0].get().strip()

            except Exception:

                disp = ""

            if not disp: continue

            excel_name = final_map.get(disp)

            if not excel_name: continue



            plist = excel_people_with_marks.get(excel_name, [])

            if not plist: continue



            for entry in plist:

                # date du marquage dans l'Excel

                try:

                    d = date.fromisoformat(entry["date"])

                except Exception:

                    continue



                # Règle demandée :

                # - si "astreinte" (BROWN/PURPLE) -> on applique le LENDEMAIN matin

                # - sinon -> on applique tel quel

                reason = (entry.get("reason") or "").strip().lower()

                if reason == "astreinte":

                    target_date = d + timedelta(days=1)

                    target_span = "MATIN"

                else:

                    target_date = d

                    target_span = (entry.get("span", "") or "").strip()



                # On n'applique que si la date cible reste dans la plage importée et dans la semaine courante

                if target_date < import_start or target_date > import_end:

                    continue

                if not (week_start <= target_date <= week_end):

                    continue



                day_idx = target_date.weekday()

                try:

                    cell_tuple = row[4 + day_idx]

                except Exception:

                    cell_tuple = None

                if not (isinstance(cell_tuple, tuple) and cell_tuple):

                    continue



                toggle = cell_tuple[0]

                try:

                    cur = toggle._var.get()

                except Exception:

                    try:

                        cur = toggle.cget("text")

                    except Exception:

                        cur = ""



                merged = _merge_span(cur, target_span)

                if hasattr(toggle, "set_state") and callable(toggle.set_state):

                    toggle.set_state(merged)

                else:

                    try:

                        toggle._var.set(merged)

                        toggle.config(text=merged)

                    except Exception:

                        pass



                # --- Nouveau : origine + log --------------------------------

                pretty_reason = {

                    "souhait": "Absence souhaitee",

                    "repos": "Repos de garde",

                    "formation": "Absence formation",

                    "astreinte": "Astreinte"

                }.get(reason, (reason or "Absence").capitalize())

                span_label = (target_span or "").strip()

                if hasattr(toggle, "get_log"):

                    try:

                        existing_note = toggle.get_log() or ""

                    except Exception:

                        existing_note = ""

                else:

                    existing_note = getattr(toggle, "log_text", "")

                note_parts = [part for part in (pretty_reason, span_label, target_date.strftime("%d/%m/%Y")) if part]

                new_note = "Import absence: " + " - ".join(note_parts)

                if existing_note:

                    note_lines = existing_note.splitlines()

                    if new_note not in note_lines:

                        combined_note = existing_note + ("\n" if existing_note.strip() else "") + new_note

                    else:

                        combined_note = existing_note

                else:

                    combined_note = new_note

                if hasattr(toggle, "set_origin"):

                    try:

                        toggle.set_origin("import_absence", log_text=combined_note, notify=False)

                    except Exception:

                        pass

                else:

                    try:

                        toggle.origin = "import_absence"

                        toggle.log_text = combined_note

                    except Exception:

                        pass

                try:

                    toggle._apply_origin_style()

                except Exception:

                    pass

                try:

                    toggle._notify_change()

                except Exception:

                    pass

            # --- FIN DU NOUVEAU BLOC ---





        try: s.update_counts()

        except Exception: pass

        try: g.update_colors(None)

        except Exception: pass

        try: g.auto_resize_all_columns()

        except Exception: pass



    # ---------- Récap ----------

    rc = Counter(m["summary"])

    details = [f"- {k} : {rc[k]}" for k in ("souhait","formation","repos","astreinte") if k in rc]

    nb_mapped  = sum(1 for v in final_map.values() if v)

    nb_ignored = sum(1 for v in final_map.values() if not v)



    msg = [

        f"{m['sheet']} {m['year']} → {m['people_all_count']} personne(s) listées (Excel), "

        f"dont {m['people_with_marks_count']} avec au moins une absence.",

        f"Semaines actives : {len(weeks)}  →  " + ", ".join(d.strftime('%d/%m') for d in weeks),

        f"Correspondances appliquées (Contraintes → Excel) : {nb_mapped}, ignorées : {nb_ignored}."

    ]

    if closed_labels:

        msg.append("Fermetures (gris) appliquées : " + ", ".join(sorted(set(closed_labels))))

    if details:

        msg.extend(["", "Détails par type (Excel) :"] + details)



    messagebox.showinfo("Import Absences", "\n".join(msg))

    setattr(root, "last_absence_import", months)



def import_conflicts_from_pkl(root, notebook, tabs_data):

    """

    Importe les CONFLITS depuis un autre planning .pkl :

      - lit les demi-journées occupées (planning A),

      - applique ces demi-journées comme ABSENCE (MATIN/AP MIDI/Journée) dans le tableau de Contraintes du planning B,

      - NE CRÉE AUCUNE LIGNE dans B,

      - reproduit la même logique de matching des noms que l'import d'absences Excel (auto + boîte de dialogue).

    """

    import tkinter as tk

    from tkinter import filedialog, messagebox, ttk

    import pickle, re, difflib

    from collections import defaultdict



    DAY_LABELS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]



    # ---------- Helpers internes (copie de la logique de l'import Excel) ----------

    STOPWORDS = {

        "DE","DU","DES","D","LE","LA","LES","DEL","DELA","DA","DOS","DO",

        "VAN","VON","DI","DELLA","MC","MAC"

    }

    INVISIBLES = {"\u00A0","\u2007","\u202F","\u200B","\u200C","\u200D","\uFEFF"}



    def _strip_invisibles(s: str) -> str:

        if not isinstance(s, str):

            return ""

        for ch in INVISIBLES:

            s = s.replace(ch, " ")

        return s



    def _norm_name(s: str) -> str:

        # _norm est défini en haut du fichier (module-level) → on le réutilise

        txt = _strip_invisibles(s)

        txt = _norm(txt).upper()

        txt = txt.replace("’", "'").replace("‐","-").replace("–","-").replace("—","-")

        txt = re.sub(r"[.\-_/(){}\[\],;:!?\t]+", " ", txt)

        txt = re.sub(r"\s+", " ", txt).strip()

        return txt



    def _tokens(s: str) -> list[str]:

        txt = _norm_name(s)

        if not txt:

            return []

        toks = []

        for token in txt.split():

            # découpe aussi "JEAN-PIERRE"

            parts = [p for p in token.split("-") if p]

            for p in parts:

                if p and p not in STOPWORDS:

                    toks.append(p)

        return toks



    def _tokens_len_ge2(s: str) -> set[str]:

        return {t for t in _tokens(s) if len(t) >= 2}



    def _token_key(s: str) -> str:

        toks = sorted(set(_tokens(s)))

        return " ".join(toks)



    def _token_set_ratio(a: str, b: str) -> float:

        A, B = set(_tokens(a)), set(_tokens(b))

        if not A or not B:

            return 0.0

        inter = len(A & B)

        return (2.0 * inter) / (len(A) + len(B))



    def _seq_ratio(a: str, b: str) -> float:

        return difflib.SequenceMatcher(None, _norm_name(a), _norm_name(b)).ratio()



    def _combined_score(a: str, b: str) -> float:

        # pondération identique à l’import Excel

        return 0.60 * _seq_ratio(a, b) + 0.40 * _token_set_ratio(a, b)



    INITIALS_AUTO_SCORE = 0.60

    INITIALS_SHORTLIST_SCORE = 0.50

    INITIALS_SHORTLIST_LIMIT = 6



    def _long_tokens(name: str) -> list[str]:

        return [t for t in _tokens(name) if len(t) >= 3]



    def _share_token_len_ge3(a: str, b: str) -> bool:

        return bool(set(_long_tokens(a)) & set(_long_tokens(b)))



    def _has_conflicting_long_tokens(a: str, b: str) -> bool:

        tokens_a = _long_tokens(a)

        tokens_b = _long_tokens(b)

        if not tokens_a or not tokens_b:

            return False

        def _matches(tok: str, pool: list[str]) -> bool:

            prefix = tok[:3]

            for other in pool:

                if tok == other or (prefix and prefix == other[:3]):

                    return True

            return False

        extra_a = [tok for tok in tokens_a if not _matches(tok, tokens_b)]

        extra_b = [tok for tok in tokens_b if not _matches(tok, tokens_a)]

        return bool(extra_a and extra_b)



    def _initial_variants(full: str) -> set[str]:

        txt = _norm_name(full)

        if not txt:

            return set()

        toks = txt.split()



        def first_letters(token: str) -> str:

            parts = [p for p in token.split("-") if p]

            return "".join(p[0] for p in parts if p)



        out = set()

        if len(toks) == 1:

            t = toks[0]

            if len(t) >= 1: out.add(t[:1])

            if len(t) >= 2: out.add(t[:2])

            return {x.replace(" ","").upper() for x in out}



        first = first_letters(toks[0])

        last  = first_letters(toks[-1])

        if first and last:

            out.add((first[:1] + last[:1]))

            out.add((last[:1] + first[:1]))

            if len(first) >= 2: out.add(first[:2] + last[:1])

            if len(last)  >= 2: out.add(first[:1] + last[:2])

        out.add((toks[0] + " " + (last[:1] if last else "")).replace(" ",""))

        if first:

            out.add((first[:1] + " " + toks[-1]).replace(" ",""))

        if first or last:

            out.add((first or "") + (last or ""))

        return {x.upper() for x in out if x}



    # _merge_span existe déjà au niveau module (on le réutilise pour fusionner M/AP → Journée)



    # ---------- 0) Sélection du .pkl source ----------

    src_path = filedialog.askopenfilename(

        title="Importer des conflits depuis un planning (.pkl)",

        defaultextension=".pkl",

        filetypes=[("Pickle Files", "*.pkl"), ("Tous fichiers", "*.*")]

    )

    if not src_path:

        return



    src_work_posts = []

    try:

        with open(src_path, "rb") as f:

            packed = pickle.load(f)

        # Format de sauvegarde courant : packed[0] = all_week_status

        all_week_status = packed[0]
        if isinstance(packed, (list, tuple)) and len(packed) >= 2:

            src_work_posts = packed[1] or []

    except Exception as e:

        messagebox.showerror("Import Conflits", f"Impossible de lire le .pkl : {e}")

        return



    # ---------- 1) Étiquettes semaines A/B ----------

    def _label_of_src_week(idx, wk):

        try:

            if len(wk) >= 6:

                return wk[4] or f"Semaine source {idx+1}"

            return wk[4] if len(wk) >= 5 and wk[4] else f"Semaine source {idx+1}"

        except Exception:

            return f"Semaine source {idx+1}"



    src_labels = [_label_of_src_week(i, wk) for i, wk in enumerate(all_week_status)]



    def _label_of_dst_week(i, g):

        try:

            return g.week_label.cget("text") or f"Semaine {i+1}"

        except Exception:

            return f"Semaine {i+1}"



    dst_labels = [_label_of_dst_week(i, tabs_data[i][0]) for i in range(len(tabs_data))]



    # --- 2) Boîte : mapping Semaine B -> Semaine A (ou Ignorer) ---

    def ask_week_mapping(parent, dst_labels, src_labels):

        import re, difflib, unicodedata

        from datetime import datetime



        def _parse_date(lbl: str):

            if not isinstance(lbl, str): return None

            m = re.search(r"(\d{2})/(\d{2})/(\d{4})", lbl)

            if not m: return None

            try:

                return datetime.strptime(m.group(0), "%d/%m/%Y").date()

            except Exception:

                return None



        def _norm_text(s: str) -> str:

            if not isinstance(s, str): return ""

            s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")

            s = s.lower()

            s = s.replace("’", "'")

            # on retire des mots peu discriminants

            for w in ("semaine", "du", "de", "la", "le", ":", "-", "_"):

                s = s.replace(w, " ")

            s = re.sub(r"\s+", " ", s).strip()

            return s



        # Pré-indexe les sources par date et par libellé normalisé

        src_dates = [_parse_date(x) for x in src_labels]

        src_norms = [_norm_text(x) for x in src_labels]



        if not dst_labels:

            return {}



        win = tk.Toplevel(parent); win.title("Associer les semaines (B → A)")

        try: win.geometry(f"+{parent.winfo_rootx()+60}+{parent.winfo_rooty()+60}")

        except Exception: pass



        frm = tk.Frame(win); frm.pack(fill="both", expand=True, padx=8, pady=8)

        tk.Label(frm, text="Semaine dans le planning B", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky="w", padx=4, pady=4)

        tk.Label(frm, text="Semaine correspondante dans le planning A", font=("Arial", 9, "bold")).grid(row=0, column=1, sticky="w", padx=4, pady=4)



        combos = []

        for i, dst_lbl in enumerate(dst_labels, start=1):

            tk.Label(frm, text=dst_lbl).grid(row=i, column=0, sticky="w", padx=4, pady=2)



            values = ["⟶ Ignorer"] + src_labels

            # ---- Détermination de la PRÉ-SÉLECTION ----

            pre_idx = None

            d_date  = _parse_date(dst_lbl)

            d_norm  = _norm_text(dst_lbl)



            # 1) match par date exacte

            if d_date is not None and d_date in src_dates:

                pre_idx = src_dates.index(d_date)

            # 2) match par libellé normalisé exact

            elif d_norm and d_norm in src_norms:

                pre_idx = src_norms.index(d_norm)

            else:

                # 3) similarité la plus haute (si très forte)

                best_j, best_r = None, 0.0

                for j, s_norm in enumerate(src_norms):

                    r = difflib.SequenceMatcher(None, d_norm, s_norm).ratio()

                    if r > best_r:

                        best_r, best_j = r, j

                if best_j is not None and best_r >= 0.90:

                    pre_idx = best_j



            default_value = "⟶ Ignorer" if pre_idx is None else values[1 + pre_idx]

            var = tk.StringVar(value=default_value)

            cb  = ttk.Combobox(frm, values=values, textvariable=var, state="readonly", width=45)

            cb.grid(row=i, column=1, sticky="w", padx=4, pady=2)

            combos.append(var)



        chosen = {}

        def _ok():

            for idx, var in enumerate(combos):  # idx = index de la semaine B

                val = var.get()

                if val == "⟶ Ignorer":

                    chosen[idx] = None

                else:

                    try:

                        chosen[idx] = src_labels.index(val)

                    except ValueError:

                        chosen[idx] = None

            win.destroy()



        tk.Button(frm, text="OK", command=_ok, width=12).grid(row=len(dst_labels)+1, column=0, columnspan=2, pady=8)

        win.transient(parent); win.grab_set(); parent.wait_window(win)

        return chosen



    # NEW — construit la correspondance de semaines B→A choisie par l'utilisateur

    week_map = ask_week_mapping(root, dst_labels, src_labels)

    if not week_map:

        return



    # ---------- 3) Référentiel des NOMS présents dans Contraintes (B) ----------

    constraint_display_names = []

    seen_norm = set()

    for (_g, c, _s) in tabs_data:

        for row in getattr(c, "rows", []):

            try:

                disp = row[0].get().strip()

            except Exception:

                disp = ""

            if not disp:

                continue

            nd = _norm_name(disp)

            if nd not in seen_norm:

                seen_norm.add(nd)

                constraint_display_names.append(disp)



    # ---------- 4) Extraire les créneaux occupés dans A, par semaine ----------

    # table_data (wk[0]) = contenu texte du planning principal A.

    # i paire = MATIN, i impaire = AP MIDI (par poste).



    per_src_week = []      # liste par semaine source : dict src_name -> {day_idx -> {(span, post_idx)}}

    all_src_names = set()



    for wk in all_week_status:

        try:

            table_data = wk[0]

        except Exception:

            table_data = []



        name_to_days = defaultdict(lambda: defaultdict(set))

        for r, row in enumerate(table_data or []):

            if row is None:

                continue

            is_morning = (r % 2 == 0)

            post_idx = r // 2

            span = "MATIN" if is_morning else "AP MIDI"

            for j, cell_txt in enumerate(row):

                if not cell_txt:

                    continue

                for token in _split_people(str(cell_txt).strip()):

                    nm = token.strip()

                    if not nm:

                        continue

                    norm = _norm_name(nm)

                    if not norm or len(norm) < 2:

                        continue

                    name_to_days[nm][j].add((span, post_idx))

                    all_src_names.add(nm)

        per_src_week.append(name_to_days)



    src_all_names = list(all_src_names)



    # ---------- 5) Index côté source pour matching ----------

    src_norm_to_names  = defaultdict(list)

    src_tokenkey_to    = defaultdict(list)

    src_initials_to    = defaultdict(set)

    src_token_to_names = defaultdict(set)



    for name in src_all_names:

        nn = _norm_name(name)

        src_norm_to_names[nn].append(name)

        tk_key = _token_key(name)

        if tk_key:

            src_tokenkey_to[tk_key].append(name)

        for iv in _initial_variants(name):

            src_initials_to[iv].add(name)

        for t in _tokens_len_ge2(name):

            src_token_to_names[t].add(name)



    # ---------- 6) Matching Contraintes(B) → Noms(source A) (auto + dialogue) ----------

    auto_map, to_confirm, to_prompt = {}, [], []



    for disp in constraint_display_names:

        nd  = _norm_name(disp)

        tkd = _token_key(disp)

        ivd = _initial_variants(disp)



        # (1) égalité normalisée stricte

        if nd in src_norm_to_names:

            cands = list(dict.fromkeys(src_norm_to_names[nd]))

            if len(cands) == 1:

                auto_map[disp] = cands[0]

            else:

                to_confirm.append((disp, cands))

            continue



        # (2) ensemble de tokens

        if tkd in src_tokenkey_to:

            cands = list(dict.fromkeys(src_tokenkey_to[tkd]))

            if len(cands) == 1:

                auto_map[disp] = cands[0]

            else:

                to_confirm.append((disp, cands))

            continue



        # (3) initiales

        hits = set()

        for token in ivd:

            if token in src_initials_to:

                hits |= src_initials_to[token]

        if hits:

            ordered_hits = sorted(hits, key=_norm_name)

            scored_hits = []

            for nm in ordered_hits:

                if not _share_token_len_ge3(disp, nm):

                    continue

                if _has_conflicting_long_tokens(disp, nm):

                    continue

                score = _combined_score(disp, nm)

                scored_hits.append((nm, score))

            if len(scored_hits) == 1 and scored_hits[0][1] >= INITIALS_AUTO_SCORE:

                auto_map[disp] = scored_hits[0][0]

                continue

            if scored_hits:

                scored_hits.sort(key=lambda x: (-x[1], _norm_name(x[0])))

                shortlist_fmt = []

                for idx, (nm, score) in enumerate(scored_hits):

                    if idx >= INITIALS_SHORTLIST_LIMIT:

                        break

                    if idx > 0 and score < INITIALS_SHORTLIST_SCORE:

                        break

                    shortlist_fmt.append(f"{nm} ({int(score*100)}%)")

                if shortlist_fmt:

                    to_confirm.append((disp, shortlist_fmt))

                    continue



        # (4) fuzzy shortlist

        cand_names = set()

        for t in _tokens_len_ge2(disp):

            cand_names |= src_token_to_names.get(t, set())

        if not cand_names:

            cand_names = set(src_all_names)



        scored = [(nm, _combined_score(disp, nm)) for nm in cand_names]

        scored.sort(key=lambda x: (-x[1], _norm_name(x[0])))



        if scored:

            top_name, top_sc = scored[0]

            second_sc = scored[1][1] if len(scored) >= 2 else 0.0

            if top_sc >= 0.92 and (top_sc - second_sc) >= 0.05:

                auto_map[disp] = top_name

            else:

                shortlist = [(nm, sc) for (nm, sc) in scored if sc >= 0.82][:7]

                if shortlist:

                    to_confirm.append((disp, [f"{nm} ({int(sc*100)}%)" for nm, sc in shortlist]))

                else:

                    to_prompt.append(disp)

        else:

            to_prompt.append(disp)



    # Boîte de validation (même esprit que l'import Excel)

    def _resolve_with_dialog(parent, items_with_scores, items_no_candidates, all_names):

        if not items_with_scores and not items_no_candidates:

            return {}

        win = tk.Toplevel(parent); win.title("Valider les correspondances (Contraintes → Planning A)")

        try:

            win.geometry(f"+{parent.winfo_rootx()+40}+{parent.winfo_rooty()+40}")

        except Exception:

            pass



        canvas = tk.Canvas(win, borderwidth=0, highlightthickness=0)

        vbar   = tk.Scrollbar(win, orient="vertical", command=canvas.yview)

        frame  = tk.Frame(canvas)

        canvas.create_window((0,0), window=frame, anchor="nw")

        canvas.configure(yscrollcommand=vbar.set)



        def _on_cfg(_e=None):

            canvas.configure(scrollregion=canvas.bbox("all"))



        frame.bind("<Configure>", _on_cfg)

        canvas.grid(row=0, column=0, sticky="nsew"); vbar.grid(row=0, column=1, sticky="ns")

        win.grid_rowconfigure(0, weight=1); win.grid_columnconfigure(0, weight=1)



        tk.Label(frame, text="Nom dans Contraintes (planning B)", font=("Arial", 9, "bold")).grid(row=0, column=0, padx=6, pady=4, sticky="w")

        tk.Label(frame, text="Nom correspondant (planning A)",   font=("Arial", 9, "bold")).grid(row=0, column=1, padx=6, pady=4, sticky="w")



        rows_vars = []

        row_idx = 1

        for disp, labeled in items_with_scores:

            tk.Label(frame, text=disp).grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")

            values = ["⟶ Ignorer"] + list(labeled)

            var = tk.StringVar(value=values[0])

            cb  = ttk.Combobox(frame, values=values, textvariable=var, width=50, state="readonly")

            cb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="w")

            rows_vars.append((disp, var))

            row_idx += 1



        for disp in items_no_candidates:

            tk.Label(frame, text=disp).grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")

            values = ["⟶ Ignorer"] + sorted(all_names, key=_norm_name)

            var = tk.StringVar(value=values[0])

            cb  = ttk.Combobox(frame, values=values, textvariable=var, width=50, state="readonly")

            cb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="w")

            rows_vars.append((disp, var))

            row_idx += 1



        chosen = {}

        def on_ok():

            for disp, var in rows_vars:

                val = var.get()

                if val == "⟶ Ignorer":

                    chosen[disp] = None

                else:

                    base = re.sub(r"\s*\(\d+%\)$", "", val).rstrip()

                    chosen[disp] = base

            win.destroy()



        btnf = tk.Frame(win); btnf.grid(row=1, column=0, columnspan=2, pady=6)

        tk.Button(btnf, text="OK", command=on_ok, width=12).pack()

        win.transient(parent); win.grab_set(); parent.wait_window(win)

        return chosen



    manual_choices = _resolve_with_dialog(root, to_confirm, to_prompt, src_all_names)



    final_map = {}

    for disp in constraint_display_names:

        if disp in auto_map:

            final_map[disp] = auto_map[disp]

        elif disp in manual_choices:

            final_map[disp] = manual_choices[disp]

        else:

            final_map[disp] = None



    # ---------- 7) Injection : on ne touche qu'au tableau Contraintes ----------

    total_marks, touched_tabs = 0, set()



    for i in range(len(tabs_data)):

        g, c, s = tabs_data[i]

        src_idx = week_map.get(i)

        if src_idx is None or src_idx < 0 or src_idx >= len(per_src_week):

            # rien à importer pour cette semaine

            try: s.update_counts()

            except Exception: pass

            try: g.update_colors(None)

            except Exception: pass

            try: g.auto_resize_all_columns()

            except Exception: pass

            continue



        name_to_days = per_src_week[src_idx]  # dict: src_name -> {day_idx -> {(span, post_idx)}}



        for row in getattr(c, "rows", []):

            try:

                disp = row[0].get().strip()

            except Exception:

                disp = ""

            if not disp:

                continue

            src_name = final_map.get(disp)

            if not src_name:

                continue



            day_map = name_to_days.get(src_name, {})

            if not day_map:

                continue



            for day_idx, entries in day_map.items():

                try:

                    cell_tuple = row[4 + int(day_idx)]

                except Exception:

                    cell_tuple = None

                if not (isinstance(cell_tuple, tuple) and cell_tuple):

                    continue



                toggle = cell_tuple[0]

                try:

                    cur = toggle._var.get()

                except Exception:

                    try:

                        cur = toggle.cget("text")

                    except Exception:

                        cur = ""



                span_values = set()
                post_indices = set()

                for span_val, post_idx in entries:

                    if span_val:

                        span_values.add(span_val)

                    if post_idx is not None:

                        post_indices.add(post_idx)

                def _span_sort_key(value: str):

                    order = {"MATIN": 0, "AP MIDI": 1, "JOURNEE": 2}

                    norm = (value or "").upper()

                    return (order.get(norm, 99), value)

                ordered_spans = sorted(span_values, key=_span_sort_key)

                merged = cur

                for sp in ordered_spans:  # ordre stable

                    merged = _merge_span(merged, sp)  # fonction module-level



                before = cur

                if hasattr(toggle, "set_state") and callable(toggle.set_state):

                    toggle.set_state(merged)

                    after = toggle._var.get()

                else:

                    try:

                        toggle._var.set(merged)

                        toggle.config(text=merged)

                    except Exception:

                        pass

                    after = merged



                if (before or "") != (after or ""):

                    total_marks += 1

                    touched_tabs.add(i)



                    # Marque l'origine + note de base pour cette absence

                    if hasattr(toggle, "get_log"):

                        try:

                            existing_note = toggle.get_log() or ""

                        except Exception:

                            existing_note = ""

                    else:

                        existing_note = getattr(toggle, "log_text", "")

                    day_label = DAY_LABELS[day_idx] if 0 <= day_idx < len(DAY_LABELS) else f"Jour {day_idx + 1}"

                    span_txt = ", ".join(ordered_spans) if ordered_spans else ""

                    post_names = []

                    for post_idx in sorted(post_indices):

                        label = ""

                        if 0 <= post_idx < len(src_work_posts):

                            label = (src_work_posts[post_idx] or "").strip()

                        if not label:

                            label = f"Poste #{post_idx + 1}"

                        post_names.append(label)

                    # conserve l'ordre d'apparition tout en supprimant les doublons

                    seen_posts = set()

                    unique_post_names = []

                    for name in post_names:

                        if name not in seen_posts:

                            seen_posts.add(name)

                            unique_post_names.append(name)

                    post_txt = ", ".join(unique_post_names) if unique_post_names else ""

                    base_parts = [

                        day_label,

                        span_txt,

                        f"Poste source: {post_txt}" if post_txt else "",

                        f"Semaine source: {src_labels[src_idx]}",

                    ]

                    base_parts = [part for part in base_parts if part]

                    base_note = "Import conflit: " + " - ".join(base_parts) if base_parts else "Import conflit"

                    if existing_note:

                        lines = existing_note.splitlines()

                        if base_note not in lines:

                            combined_note = existing_note + ("\n" if existing_note.strip() else "") + base_note

                        else:

                            combined_note = existing_note

                    else:

                        combined_note = base_note

                    if hasattr(toggle, "set_origin"):

                        try:

                            toggle.set_origin("import_conflict", log_text=combined_note, notify=False)

                        except Exception:

                            pass

                    else:

                        try:

                            toggle.origin = "import_conflict"

                            toggle.log_text = combined_note

                        except Exception:

                            pass

                    try:

                        toggle._apply_origin_style()

                    except Exception:

                        pass

                    try:

                        toggle._notify_change()

                    except Exception:

                        pass



        # rafraîchis la semaine i

        try: s.update_counts()

        except Exception: pass

        try: g.update_colors(None)

        except Exception: pass

        try: g.auto_resize_all_columns()

        except Exception: pass



    messagebox.showinfo("Import Conflits", f"Import terminé.\nAbsences ajoutées/modifiées : {total_marks}")



def check_cross_planning_conflicts_from_pkl(root, notebook, tabs_data):

    """

    Analyse un planning .pkl (A) et détecte les conflits avec le planning B (ouvert) :

      - même personne, même semaine mappée, même jour, même demi-journée (MATIN/AP MIDI)

      - marquage visuel discret dans B (badge <->, sans changer les couleurs existantes)

      - fenêtre de rapport cliquable

    Matching des noms : strict, n'affiche la boîte que pour les DOUTES réels.

    """

    import tkinter as tk

    from tkinter import ttk, filedialog, messagebox

    import pickle, re, difflib

    from collections import defaultdict



    # ---------- Helpers matching (logique alignée sur import absences) ----------

    STOPWORDS = {"DE","DU","DES","D","LE","LA","LES","DEL","DELA","DA","DOS","DO","VAN","VON","DI","DELLA","MC","MAC"}

    INVISIBLES = {"\u00A0","\u2007","\u202F","\u200B","\u200C","\u200D","\uFEFF"}



    def _strip_invisibles(s: str) -> str:

        if not isinstance(s, str): return ""

        for ch in INVISIBLES: s = s.replace(ch, " ")

        return s



    # Réutilise _norm du module si présent

    def _norm_fallback(s: str) -> str:

        try:

            return _norm(s)

        except Exception:

            import unicodedata

            s = _strip_invisibles(s or "")

            s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")

            return s



    def _norm_name(s: str) -> str:

        s = _strip_invisibles(s)

        s = _norm_fallback(s).upper()

        s = s.replace("’", "'").replace("‐","-").replace("–","-").replace("—","-")

        s = re.sub(r"[.\-_/(){}\[\],;:!?\t]+", " ", s)

        s = re.sub(r"\s+", " ", s).strip()

        return s



    def _tokens(s: str) -> list[str]:

        txt = _norm_name(s)

        if not txt: return []

        toks = []

        for token in txt.split():

            parts = [p for p in token.split("-") if p]

            for p in parts:

                if p and p not in STOPWORDS:

                    toks.append(p)

        return toks



    def _tokens_len_ge2(s: str) -> set[str]:

        return {t for t in _tokens(s) if len(t) >= 2}



    def _token_key(s: str) -> str:

        toks = sorted(set(_tokens(s)))

        return " ".join(toks)



    def _token_set_ratio(a: str, b: str) -> float:

        A, B = set(_tokens(a)), set(_tokens(b))

        if not A or not B: return 0.0

        inter = len(A & B)

        return (2.0 * inter) / (len(A) + len(B))



    def _seq_ratio(a: str, b: str) -> float:

        return difflib.SequenceMatcher(None, _norm_name(a), _norm_name(b)).ratio()



    def _combined_score(a: str, b: str) -> float:

        # mix séquence + recouvrement de tokens

        return 0.60 * _seq_ratio(a, b) + 0.40 * _token_set_ratio(a, b)



    def _initial_variants(full: str) -> set[str]:

        txt = _norm_name(full)

        if not txt: return set()

        toks = txt.split()

        def first_letters(token: str) -> str:

            parts = [p for p in token.split("-") if p]

            return "".join(p[0] for p in parts if p)

        out = set()

        if len(toks) == 1:

            t = toks[0]

            if len(t) >= 1: out.add(t[:1])

            if len(t) >= 2: out.add(t[:2])

            return {x.replace(" ","").upper() for x in out}

        first = first_letters(toks[0]); last = first_letters(toks[-1])

        if first and last:

            out.update({first[:1]+last[:1], last[:1]+first[:1]})

            if len(first) >= 2: out.add(first[:2] + last[:1])

            if len(last)  >= 2: out.add(first[:1] + last[:2])

        out.add((toks[0] + " " + (last[:1] if last else "")).replace(" ",""))

        if first: out.add((first[:1] + " " + toks[-1]).replace(" ",""))

        if first or last: out.add((first or "") + (last or ""))

        return {x.upper() for x in out if x}



    # ---------- Sélection du planning A ----------

    src_path = filedialog.askopenfilename(

        title="Vérifier Conflits inter-plannings (.pkl)",

        defaultextension=".pkl",

        filetypes=[("Fichiers Pickle", "*.pkl"), ("Tous fichiers", "*.*")]

    )

    if not src_path:

        return



    try:

        with open(src_path, "rb") as f:

            packed = pickle.load(f)

        all_week_status = packed[0]       # [(table_data, cell_av, constraints, schedule, week_label, [excluded]), ...]

        src_work_posts  = packed[1] if len(packed) >= 2 else []

    except Exception as e:

        messagebox.showerror("Conflits inter-plannings", f"Impossible de lire le .pkl : {e}")

        return



    # ---------- Libellés semaines ----------

    def _label_of_src_week(idx, wk):

        try:

            if len(wk) >= 6:

                return wk[4] or f"Semaine source {idx+1}"

            return wk[4] if len(wk) >= 5 and wk[4] else f"Semaine source {idx+1}"

        except Exception:

            return f"Semaine source {idx+1}"



    src_labels = [_label_of_src_week(i, wk) for i, wk in enumerate(all_week_status)]

    dst_labels = []

    for i, (g, _c, _s) in enumerate(tabs_data):

        try:

            dst_labels.append(g.week_label.cget("text") or f"Semaine {i+1}")

        except Exception:

            dst_labels.append(f"Semaine {i+1}")



    # ---------- Mapping B → A (semaines) avec pré-sélection intelligente ----------

    def ask_week_mapping(parent, dst_labels, src_labels):

        import re, difflib, unicodedata

        from datetime import datetime



        def _parse_date(lbl: str):

            if not isinstance(lbl, str): return None

            m = re.search(r"(\d{2})/(\d{2})/(\d{4})", lbl)

            if not m: return None

            try: return datetime.strptime(m.group(0), "%d/%m/%Y").date()

            except Exception: return None



        def _norm_text(s: str) -> str:

            if not isinstance(s, str): return ""

            s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")

            s = s.lower().replace("’","'")

            for w in ("semaine","du","de","la","le",":","-","_"): s = s.replace(w, " ")

            return re.sub(r"\s+"," ",s).strip()



        src_dates = [_parse_date(x) for x in src_labels]

        src_norms = [_norm_text(x) for x in src_labels]



        if not dst_labels: return {}

        win = tk.Toplevel(parent); win.title("Associer les semaines (B → A)")

        try: win.geometry(f"+{parent.winfo_rootx()+60}+{parent.winfo_rooty()+60}")

        except Exception: pass



        frm = tk.Frame(win); frm.pack(fill="both", expand=True, padx=8, pady=8)

        tk.Label(frm, text="Semaine B", font=("Arial", 9, "bold")).grid(row=0, column=0, sticky="w", padx=4, pady=4)

        tk.Label(frm, text="Semaine A", font=("Arial", 9, "bold")).grid(row=0, column=1, sticky="w", padx=4, pady=4)



        combos = []

        for i, dst_lbl in enumerate(dst_labels, start=1):

            tk.Label(frm, text=dst_lbl).grid(row=i, column=0, sticky="w", padx=4, pady=2)

            values = ["⟶ Ignorer"] + src_labels



            pre_idx = None

            d_date  = _parse_date(dst_lbl)

            d_norm  = _norm_text(dst_lbl)

            if d_date is not None and d_date in src_dates:

                pre_idx = src_dates.index(d_date)

            elif d_norm and d_norm in src_norms:

                pre_idx = src_norms.index(d_norm)

            else:

                best_j, best_r = None, 0.0

                for j, s_norm in enumerate(src_norms):

                    r = difflib.SequenceMatcher(None, d_norm, s_norm).ratio()

                    if r > best_r: best_r, best_j = r, j

                if best_j is not None and best_r >= 0.90:

                    pre_idx = best_j



            default_value = "⟶ Ignorer" if pre_idx is None else values[1 + pre_idx]

            var = tk.StringVar(value=default_value)

            cb  = ttk.Combobox(frm, values=values, textvariable=var, state="readonly", width=45)

            cb.grid(row=i, column=1, sticky="w", padx=4, pady=2)

            combos.append(var)



        chosen = {}

        def _ok():

            for idx, var in enumerate(combos):

                val = var.get()

                if val == "⟶ Ignorer":

                    chosen[idx] = None

                else:

                    try: chosen[idx] = src_labels.index(val)

                    except ValueError: chosen[idx] = None

            win.destroy()



        tk.Button(frm, text="OK", command=_ok, width=12).grid(row=len(dst_labels)+1, column=0, columnspan=2, pady=8)

        win.transient(parent); win.grab_set(); parent.wait_window(win)

        return chosen



    week_map = ask_week_mapping(root, dst_labels, src_labels)

    if not week_map: return



    # ---------- Récup noms A ----------

    src_all_names = set()

    for wk in all_week_status:

        table_data = wk[0] if len(wk) >= 1 else []

        for r, row in enumerate(table_data or []):

            for j, cell in enumerate(row or []):

                if not cell: continue

                for token in _split_people(str(cell).strip()):

                    nm = token.strip()

                    if not nm:

                        continue

                    norm = _norm_name(nm)

                    if not norm or len(norm) < 2:

                        continue

                    src_all_names.add(nm)

    src_all_names = sorted(src_all_names, key=_norm_name)



    # ---------- Récup noms B (tableau principal) ----------

    b_all_names = set()

    for (g, _c, _s) in tabs_data:

        for r, row in enumerate(g.table_entries):

            for j, entry in enumerate(row):

                txt = entry.get().strip()

                if not txt: continue

                for token in _split_people(txt):

                    nm = token.strip()

                    if not nm:

                        continue

                    norm = _norm_name(nm)

                    if not norm or len(norm) < 2:

                        continue

                    b_all_names.add(nm)

    b_all_names = sorted(b_all_names, key=_norm_name)



    # ---------- Index source A pour matching ----------

    src_norm_to_names  = defaultdict(list)

    src_tokenkey_to    = defaultdict(list)

    src_initials_to    = defaultdict(set)

    src_token_to_names = defaultdict(set)



    for name in src_all_names:

        nn = _norm_name(name)

        src_norm_to_names[nn].append(name)

        tk_key = _token_key(name)

        if tk_key: src_tokenkey_to[tk_key].append(name)

        for iv in _initial_variants(name): src_initials_to[iv].add(name)

        for t in _tokens_len_ge2(name):   src_token_to_names[t].add(name)



    # ---------- Matching B → A (auto + DOUTES réels uniquement) ----------

    AUTO_STRONG = 0.965    # auto-match si très proche

    ASK_MIN     = 0.92     # on ne questionne que si top >= 0.92

    GAP_MAX     = 0.03     # et si les deux meilleurs sont très proches (vrai doute)

    SHORTLIST_K = 5        # max propositions en boîte



    auto_map: dict[str, str] = {}

    to_confirm: list[tuple[str, list[str]]] = []



    for disp in b_all_names:

        nd, tkd, ivd = _norm_name(disp), _token_key(disp), _initial_variants(disp)



        # (1) égalité normalisée stricte

        if nd in src_norm_to_names:

            cands = list(dict.fromkeys(src_norm_to_names[nd]))

            if len(cands) == 1:

                auto_map[disp] = cands[0]

            else:

                # vrai doute seulement si les libellés sont distincts

                to_confirm.append((disp, cands[:SHORTLIST_K]))

            continue



        # (2) ensemble de tokens

        if tkd in src_tokenkey_to:

            cands = list(dict.fromkeys(src_tokenkey_to[tkd]))

            if len(cands) == 1:

                auto_map[disp] = cands[0]

            else:

                to_confirm.append((disp, cands[:SHORTLIST_K]))

            continue



        # (3) initiales

        hits = set()

        for token in ivd:

            if token in src_initials_to:

                hits |= src_initials_to[token]



        if hits:

            disp_tokens_3 = {t for t in _tokens(disp) if len(t) >= 3}

            def _has_shared_token(candidate: str) -> bool:

                cand_tokens = {t for t in _tokens(candidate) if len(t) >= 3}

                return bool(disp_tokens_3 & cand_tokens)



            hits_with_overlap = [nm for nm in hits if _has_shared_token(nm)]



            if len(hits) == 1 and hits_with_overlap:

                auto_map[disp] = hits_with_overlap[0]

                continue



            shortlist = sorted(hits_with_overlap or hits)[:SHORTLIST_K]

            to_confirm.append((disp, shortlist))

            continue



        # (4) fuzzy : limiter les candidats aux recouvrements plausibles

        cand_names = set()

        for t in _tokens_len_ge2(disp):

            cand_names |= src_token_to_names.get(t, set())

        for token in ivd:

            cand_names |= src_initials_to.get(token, set())



        if not cand_names:

            # pas de candidat plausible -> on IGNORE ce nom (aucune boîte)

            continue



        scored = [(nm, _combined_score(disp, nm)) for nm in cand_names]

        scored.sort(key=lambda x: (-x[1], _norm_name(x[0])))



        top_sc = scored[0][1]

        second_sc = scored[1][1] if len(scored) >= 2 else 0.0

        top_name  = scored[0][0]



        if top_sc >= AUTO_STRONG and (top_sc - second_sc) >= 0.04:

            auto_map[disp] = top_name

        elif top_sc >= ASK_MIN and (top_sc - second_sc) <= GAP_MAX:

            # vrai doute -> on propose une courte shortlist

            shortlist = [f"{nm} ({int(sc*100)}%)" for (nm, sc) in scored if sc >= ASK_MIN][:SHORTLIST_K]

            if len(shortlist) >= 2:  # afficher seulement si plusieurs proches

                to_confirm.append((disp, shortlist))

            else:

                # un seul bon candidat -> auto

                auto_map[disp] = top_name

        else:

            # correspondance trop faible -> IGNORER silencieusement

            pass



    # Boîte de validation : uniquement pour les doutes (pas de "liste complète")

    def _resolve_with_dialog(parent, items_with_scores):

        if not items_with_scores: return {}

        win = tk.Toplevel(parent); win.title("Valider correspondances (B → A)")

        try: win.geometry(f"+{parent.winfo_rootx()+40}+{parent.winfo_rooty()+40}")

        except Exception: pass

        canvas = tk.Canvas(win, borderwidth=0, highlightthickness=0)

        vbar   = tk.Scrollbar(win, orient="vertical", command=canvas.yview)

        frame  = tk.Frame(canvas)

        canvas.create_window((0,0), window=frame, anchor="nw")

        canvas.configure(yscrollcommand=vbar.set)

        def _on_cfg(_e=None): canvas.configure(scrollregion=canvas.bbox("all"))

        frame.bind("<Configure>", _on_cfg)

        canvas.grid(row=0, column=0, sticky="nsew"); vbar.grid(row=0, column=1, sticky="ns")

        win.grid_rowconfigure(0, weight=1); win.grid_columnconfigure(0, weight=1)

        tk.Label(frame, text="Nom (planning B)", font=("Arial", 9, "bold")).grid(row=0, column=0, padx=6, pady=4, sticky="w")

        tk.Label(frame, text="Nom correspondant (planning A)", font=("Arial", 9, "bold")).grid(row=0, column=1, padx=6, pady=4, sticky="w")

        rows_vars = []; row_idx = 1

        for disp, labeled in items_with_scores:

            tk.Label(frame, text=disp).grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")

            values = ["⟶ Ignorer"] + list(labeled)

            var = tk.StringVar(value=values[0])

            cb  = ttk.Combobox(frame, values=values, textvariable=var, width=50, state="readonly")

            cb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="w")

            rows_vars.append((disp, var)); row_idx += 1

        chosen = {}

        def on_ok():

            for disp, var in rows_vars:

                val = var.get()

                if val == "⟶ Ignorer":

                    chosen[disp] = None

                else:

                    chosen[disp] = re.sub(r"\s*\(\d+%\)$", "", val).rstrip()

            win.destroy()

        btnf = tk.Frame(win); btnf.grid(row=1, column=0, columnspan=2, pady=6)

        tk.Button(btnf, text="OK", command=on_ok, width=12).pack()

        win.transient(parent); win.grab_set(); parent.wait_window(win)

        return chosen



    manual_choices = _resolve_with_dialog(root, to_confirm)



    # final_map : autos + choix manuels (le reste = ignoré)

    final_map = {}

    for disp in b_all_names:

        if disp in auto_map:

            final_map[disp] = auto_map[disp]

        elif disp in manual_choices:

            final_map[disp] = manual_choices[disp]

        else:

            final_map[disp] = None



    # Limiter aux personnes réellement bi-appartenantes (présentes dans A ET B après mapping)

    mapped_in_b = {nmB for nmB, nmA in final_map.items() if nmA}

    if not mapped_in_b:

        messagebox.showinfo("Conflits inter-plannings", "Aucune personne bi-appartenante détectée (après filtrage strict).")

        return



    # ---------- Index par semaine côté A ----------

    def half_from_row(r: int) -> str:

        return "MATIN" if (r % 2 == 0) else "AP MIDI"



    src_index = []

    for w_idx, wk in enumerate(all_week_status):

        table_data = wk[0] if len(wk) >= 1 else []

        per_name = defaultdict(list)  # normA -> [(day, half, post_idx_src)]

        for r, row in enumerate(table_data or []):

            for j, cell in enumerate(row or []):

                if not cell: continue

                for token in _split_people(str(cell).strip()):

                    nm = token.strip()

                    if not nm:

                        continue

                    norm = _norm_name(nm)

                    if not norm or len(norm) < 2:

                        continue

                    per_name[norm].append((j, half_from_row(r), r // 2))

        src_index.append(per_name)



    # ---------- Détection + marquage (badge uniquement) ----------

    DAYS = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]



    # Nettoyage d’anciens marquages

    for (g, _c, _s) in tabs_data:

        try: g.clear_cross_conflict_marks()

        except Exception: pass



    try:

        from Full_GUI import work_posts as dst_work_posts

    except Exception:

        dst_work_posts = []



    collected = []  # (tab_idx, r, j, nmB, nmA, half, postB, postA, labelB, labelA)

    seen_cell = set()



    for tab_idx, (g, _c, _s) in enumerate(tabs_data):

        src_idx = week_map.get(tab_idx)

        if src_idx is None or src_idx < 0 or src_idx >= len(src_index):

            continue

        labelB = dst_labels[tab_idx]

        labelA = src_labels[src_idx]

        indexA = src_index[src_idx]



        for r, row in enumerate(g.table_entries):

            halfB = half_from_row(r)

            postB = dst_work_posts[r // 2] if (r // 2) < len(dst_work_posts) else ""

            for j, entry in enumerate(row):

                txt = entry.get().strip()

                if not txt: continue

                has_conflict_here = False

                for token in _split_people(txt):

                    nmB = token.strip()

                    if not nmB or nmB not in mapped_in_b:

                        continue

                    nmA = final_map.get(nmB)

                    if not nmA: continue

                    for (dA, halfA, post_idx_src) in indexA.get(_norm_name(nmA), []):

                        if dA == j and halfA == halfB:

                            postA = src_work_posts[post_idx_src] if post_idx_src < len(src_work_posts) else ""

                            collected.append((tab_idx, r, j, nmB, nmA, halfB, postB, postA, labelB, labelA))

                            has_conflict_here = True

                            break

                    if has_conflict_here:

                        break

                if has_conflict_here:

                    key = (tab_idx, r, j)

                    if key not in seen_cell:

                        seen_cell.add(key)

                        try: g.mark_cross_conflict(r, j)  # badge <-> seulement

                        except Exception: pass



    # Ajout des details des conflits dans les journaux d'absence

    for (tab_idx, _r, j, nmB, _nmA, halfB, postB, postA, _labelB, labelA) in collected:

        try:

            constraints_app = tabs_data[tab_idx][1]

        except Exception:

            continue

        target_row = None

        for row in getattr(constraints_app, "rows", []):

            try:

                label = row[0].get().strip()

            except Exception:

                label = ""

            if label == nmB:

                target_row = row

                break

        if target_row is None:

            continue

        try:

            cell_tuple = target_row[4 + j]

        except Exception:

            continue

        if not (isinstance(cell_tuple, tuple) and cell_tuple):

            continue

        toggle = cell_tuple[0]

        if hasattr(toggle, "get_log"):

            try:

                existing_note = toggle.get_log() or ""

            except Exception:

                existing_note = ""

        else:

            existing_note = getattr(toggle, "log_text", "")

        lines = existing_note.splitlines() if existing_note else []

        day_label = DAY_LABELS[j] if 0 <= j < len(DAY_LABELS) else f"Jour {j + 1}"

        summary_parts = [day_label]

        if halfB:

            summary_parts.append(halfB)

        if postB:

            summary_parts.append(f"Poste B: {postB}")

        if postA:

            summary_parts.append(f"Poste source: {postA}")

        summary_parts.append(f"Semaine source: {labelA}")

        summary_line = "Import conflit: " + " - ".join(summary_parts)

        summary_idx = next((idx for idx, line in enumerate(lines) if line.startswith("Import conflit")), None)

        if summary_idx is None:

            lines.insert(0, summary_line)

        else:

            lines[summary_idx] = summary_line



        detail_parts = [f"Jour: {day_label}", f"Demi-journee: {halfB}"]

        if postB:

            detail_parts.append(f"Poste B: {postB}")

        if postA:

            detail_parts.append(f"Poste source: {postA}")

        detail_parts.append(f"Semaine source: {labelA}")

        detail_line = "Import conflit detail: " + " | ".join(detail_parts)

        if detail_line not in lines:

            lines.append(detail_line)



        combined_note = "\n".join(lines)

        if hasattr(toggle, "set_origin"):

            try:

                toggle.set_origin("import_conflict", log_text=combined_note, notify=False)

            except Exception:

                pass

        else:

            try:

                toggle.origin = "import_conflict"

                toggle.log_text = combined_note

            except Exception:

                pass

        try:

            toggle._apply_origin_style()

        except Exception:

            pass

        try:

            toggle._notify_change()

        except Exception:

            pass



    # ---------- Rapport cliquable ----------

    def show_report(parent, rows):

        win = tk.Toplevel(parent)

        win.title("Conflits inter-plannings (A ↔ B)")

        try: win.geometry(f"+{parent.winfo_rootx()+80}+{parent.winfo_rooty()+80}")

        except Exception: pass



        cols = ("semaineB","jour","demi","personne","posteB","posteA","semaineA")

        tree = ttk.Treeview(win, columns=cols, show="headings", height=16)

        headers = {

            "semaineB": "Semaine B",

            "jour":     "Jour",

            "demi":     "Demi-journée",

            "personne": "Personne",

            "posteB":   "Poste B",

            "posteA":   "Poste A",

            "semaineA": "Semaine A"

        }

        for key in cols:

            tree.heading(key, text=headers[key]); tree.column(key, width=140, stretch=True)

        tree.pack(fill="both", expand=True, padx=6, pady=6)



        item_map = {}

        for (tab_idx, r, j, nmB, nmA, halfB, postB, postA, labelB, labelA) in rows:

            iid = tree.insert("", "end", values=(labelB, DAYS[j], halfB, nmB, postB, postA, labelA))

            item_map[iid] = (tab_idx, r, j)



        def goto_selected(_e=None):

            sel = tree.selection()

            if not sel: return

            iid = sel[0]

            tab_idx, r, j = item_map.get(iid, (None, None, None))

            if tab_idx is None: return

            try: notebook.select(tab_idx)

            except Exception: pass

            try:

                g = tabs_data[tab_idx][0]

                g.scroll_to_cell(r, j)

            except Exception: pass



        tree.bind("<Double-1>", goto_selected)



        btns = tk.Frame(win); btns.pack(fill="x", padx=6, pady=4)

        tk.Button(btns, text="Aller à la cellule", command=goto_selected, width=18).pack(side="left")

        def clear_marks():

            for (g, _c, _s) in tabs_data:

                try: g.clear_cross_conflict_marks()

                except Exception: pass

        tk.Button(btns, text="Effacer les marquages", command=clear_marks, width=20).pack(side="right")



        if not rows:

            tk.Label(win, text="Aucun conflit détecté.", fg="gray").pack(pady=4)



    show_report(root, collected)

    messagebox.showinfo("Conflits inter-plannings", f"Analyse terminée.\nConflits détectés : {len(collected)}")

