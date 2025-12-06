from __future__ import annotations

DEFAULT_DAY_COLUMN_WIDTH = 15
from openpyxl.utils import get_column_letter

def register_day_column_width(width_map, column_index, value, base_width=DEFAULT_DAY_COLUMN_WIDTH, max_width=60):
    if value is None:
        return
    text = str(value)
    if not text.strip():
        return
    max_line = max(len(line) for line in text.splitlines())
    target = max(base_width, min(max_width, max_line + 2))
    letter = get_column_letter(column_index)
    current = width_map.get(letter, base_width)
    if target > current:
        width_map[letter] = target





def export_to_excel(root, tabs_data, days, work_posts, POST_INFO):
    """
    Exporte chaque semaine (onglet) dans une feuille Excel :
      - Planning principal (horaires + initiales)
      - PDS (astreintes)
      - Absence (texte des Ã©tats)
      - Tableau de dÃ©compte (ShiftCountTable)
      - Statistiques individuelles par PERSONNE :
          * Comptes PAR POSTE EXACT (colonne A), pas par grande catÃ©gorie
          * Vac vend aprem (Oui/Non)
          * Double vacations (>= 2 postes dans le mÃªme crÃ©neau, mÃªme jour)
          * Scanner toute la Journée (au moins 1 poste Scanner matin + au moins 1 poste Scanner aprem le mÃªme jour)
          * Moyenne de vacs par semaine (sur lâ€™ensemble des onglets) en comptant les doubles vacations comme 2
      - Graphique en â€œpizzaâ€ (rÃ©partition par poste) aux couleurs du planning
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from tkinter import messagebox, filedialog

    # ---------- Helpers ----------
    def is_scanner_post(post_name: str) -> bool:
        s = (post_name or "").lower()
        return "scanner" in s or " ct" in s or s.endswith(" ct") or " ct-" in s

    def find_friday_index(days_list):
        for i, d in enumerate(days_list):
            if str(d).strip().lower().startswith("ven"):
                return i
        return -1

    def fmt_avg(x: float, weeks_count: int):
        if weeks_count <= 0:
            return 0
        val = x / weeks_count
        return int(val) if abs(val - int(val)) < 1e-9 else round(val, 1)

    def norm_person(s: str) -> str:
        """Normalise un identifiant personne (espaces/maj) pour la moyenne multi-onglets."""
        return " ".join(str(s or "").strip().split()).upper()

    # --- Choix du fichier ---
    file_path = filedialog.asksaveasfilename(
        parent=root,
        title="Exporter le planning",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not file_path:
        return None

    # ===========================================================
    # PRE-PASS: weekly totals per person across all weeks
    #           (driven by the shift-count table)
    # ===========================================================
    from collections import defaultdict, Counter
    import re

    _MULTI_NAME_SPLIT_RE = re.compile(r"[\n,;/&+]+")

    def _parse_cell_names(raw_text, valid_names=None):
        text = str(raw_text or '').strip()
        if not text or text.lower() == 'x':
            return []
        normalized = re.sub(r"\s{2,}", " ", text)
        parts = [p.strip() for p in _MULTI_NAME_SPLIT_RE.split(normalized) if p.strip()]
        if not parts:
            parts = [text]
        if not valid_names:
            return list(dict.fromkeys(parts))
        norm_map = {norm_person(name): name for name in valid_names}
        seen = set()
        result = []
        for part in parts:
            key = norm_person(part)
            if key in norm_map and key not in seen:
                result.append(norm_map[key])
                seen.add(key)
        if result:
            return result
        upper_text = ' '.join(text.upper().split())
        for key, original in norm_map.items():
            pattern = r"(?<!\S)" + re.escape(key) + r"(?!\S)"
            if re.search(pattern, upper_text) and key not in seen:
                result.append(original)
                seen.add(key)
        return result

    n_weeks   = len(tabs_data)
    day_count_default = len(days)
    # Key = normalized person name -> [weekly totals]
    weekly_totals_by_person = defaultdict(lambda: [0] * n_weeks)

    for w_idx, (gui_w, constraints_w, shift_table_w) in enumerate(tabs_data):
        tree = getattr(shift_table_w, 'tree', None)
        columns = list(getattr(shift_table_w, 'columns', []) or [])
        total_idx = None
        for ci, col in enumerate(columns):
            if str(col).strip().lower() == 'total':
                total_idx = ci
                break
        seen_norms = set()
        norm_map = {}
        if tree is not None and total_idx is not None:
            for item in tree.get_children():
                vals = tree.item(item, 'values')
                if not vals or len(vals) <= total_idx:
                    continue
                raw_name = vals[0]
                name = str(raw_name).strip()
                if not name:
                    continue
                norm_name = norm_person(name)
                norm_map.setdefault(norm_name, name)
                raw_total = vals[total_idx]
                total = 0
                if raw_total not in (None, ''):
                    try:
                        total = float(raw_total)
                    except (TypeError, ValueError):
                        match = re.search(r"[-+]?\d+(?:[\.,]\d+)?", str(raw_total))
                        if match:
                            try:
                                total = float(match.group(0).replace(',', '.'))
                            except ValueError:
                                total = 0
                        else:
                            total = 0
                if abs(total - int(total)) < 1e-9:
                    total = int(total)
                weekly_totals_by_person[norm_name][w_idx] += total
                seen_norms.add(norm_name)

        parser_valids = set(norm_map.values())
        rows = getattr(constraints_w, 'rows', None)
        if rows:
            for row in rows:
                try:
                    init = row[0].get().strip()
                except Exception:
                    init = ''
                if init:
                    parser_valids.add(init)

        table_entries = getattr(gui_w, 'table_entries', []) or []
        day_count_week = min(day_count_default, len(table_entries))
        excluded_cells = getattr(gui_w, 'excluded_from_count', set()) or set()
        for di in range(day_count_week):
            row_entries = table_entries[di]
            for p_index, _post in enumerate(work_posts):
                if p_index >= len(row_entries):
                    continue
                if (di, p_index) in excluded_cells:
                    continue
                cell = row_entries[p_index]
                if not cell:
                    continue
                try:
                    raw_value = cell.get()
                except Exception:
                    raw_value = ''
                names = _parse_cell_names(raw_value, parser_valids)
                if not names:
                    continue
                cell_seen = set()
                for name in names:
                    if not name:
                        continue
                    norm_name = norm_person(name)
                    if norm_name in cell_seen or norm_name in seen_norms:
                        continue
                    weekly_totals_by_person[norm_name][w_idx] += 1
                    cell_seen.add(norm_name)

    # --- Workbook / styles ---
    workbook = openpyxl.Workbook()
    thin_border    = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    dark_grey_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
    header_fill    = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Palette couleur PAR POSTE (fallback gris)
    post_color_map = {}
    for post in work_posts:
        col = (POST_INFO.get(post, {}).get("color") or "").lstrip("#").upper()
        post_color_map[post] = col if col else "A5A5A5"

    # --- Boucle semaines / onglets ---
    for idx, (gui_instance, constraints_instance, shift_count_table) in enumerate(tabs_data):
        if idx == 0:
            sheet = workbook.active
            sheet.title = f"Semaine {idx+1}"
        else:
            sheet = workbook.create_sheet(title=f"Semaine {idx+1}")

        table_entries = getattr(gui_instance, 'table_entries', []) or []
        day_count = min(day_count_default, len(table_entries))

        day_column_widths = {}

        # ---------- Collecte PDS ----------
        pds_by_day = [[] for _ in range(day_count)]
        rows_constraints = getattr(constraints_instance, "rows", []) or []
        for row in rows_constraints:
            if not row or len(row) <= 4:
                continue
            try:
                initials = row[0].get().strip()
            except Exception:
                initials = ""
            max_di = min(day_count, len(row) - 4)
            for di in range(max_di):
                try:
                    tpl = row[4 + di]
                except Exception:
                    continue
                if isinstance(tpl, tuple) and len(tpl) == 3:
                    try:
                        if tpl[2].get() == 1:
                            pds_by_day[di].append(initials)
                    except Exception:
                        pass

        # ---------- Titre ----------
        semaine_label = gui_instance.week_label.cget("text").strip() or f"Semaine {idx+1}"
        sheet.cell(row=1, column=1, value=semaine_label).border = thin_border

        # Layout : lignes = jours, colonnes = postes
        col_day_label = 1
        col_post_start = 2
        header_row    = 2

        # En-tÃªtes colonnes (postes)
        sheet.cell(row=header_row, column=col_day_label, value="Jour").border = thin_border
        for p_index, post in enumerate(work_posts):
            col = col_post_start + p_index
            hexcol = (POST_INFO.get(post, {}).get("color", "#DDDDDD") or "#DDDDDD").lstrip("#")
            post_fill = PatternFill(start_color=hexcol, end_color=hexcol, fill_type="solid")
            c = sheet.cell(row=header_row, column=col, value=post)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border
            c.font      = Font(bold=True)
            c.fill      = post_fill
            register_day_column_width(day_column_widths, col, post)

        # Lignes de jours
        weekend_rows = getattr(gui_instance, "weekend_rows", set()) or set()
        holiday_rows = getattr(gui_instance, "holiday_rows", set()) or set()
        holiday_dates = getattr(gui_instance, "holiday_dates", set()) or set()
        current_year = getattr(gui_instance, "current_year", None)
        current_month = getattr(gui_instance, "current_month", None)

        WEEKEND_FILL = PatternFill(start_color="FFF6F2", end_color="FFF6F2", fill_type="solid")
        HOLIDAY_FILL = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")
        WHITE_FILL   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        row_offset = header_row + 1
        for di in range(day_count):
            excel_row = row_offset + di
            try:
                lbl_text = gui_instance.day_labels[di].cget("text").strip()
            except Exception:
                lbl_text = ""
            if not lbl_text:
                lbl_text = days[di] if di < len(days) else str(di + 1)
            sheet.cell(row=excel_row, column=col_day_label, value=lbl_text).border = thin_border

            # Eval weekend/holiday
            is_weekend = di in weekend_rows
            is_holiday = di in holiday_rows
            try:
                if current_year and current_month and lbl_text.isdigit():
                    dt = date(current_year, current_month, int(lbl_text))
                    is_weekend = is_weekend or dt.weekday() >= 5
                    is_holiday = is_holiday or (dt in holiday_dates)
            except Exception:
                pass
            base_fill = HOLIDAY_FILL if is_holiday else (WEEKEND_FILL if is_weekend else WHITE_FILL)

            if di >= len(gui_instance.table_entries):
                continue
            row_entries = gui_instance.table_entries[di]
            for p_index, post in enumerate(work_posts):
                if p_index >= len(row_entries):
                    continue
                col = col_post_start + p_index
                entry = row_entries[p_index]
                try:
                    txt_init = entry.get().strip()
                except Exception:
                    txt_init = ""
                is_disabled = not gui_instance.cell_availability.get((di, p_index), True)
                cell = sheet.cell(row=excel_row, column=col, value=txt_init)
                register_day_column_width(day_column_widths, col, txt_init)
                if is_disabled:
                    cell.fill = dark_grey_fill
                    cell.font = Font(color="999999")
                else:
                    cell.fill = base_fill
                    if txt_init:
                        cell.font = Font(bold=True, color="FF0000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = thin_border

        # Largeur colonnes
        max_col = col_post_start + len(work_posts) - 1
        for c in range(1, max_col+1):
            let = get_column_letter(c)
            sheet.column_dimensions[let].width = sheet.column_dimensions.get(let, type('', (), {'width':0})()).width or (22 if c == 1 else 14)

        # ---------- Tableau de dÃ©compte (ShiftCountTable) ----------
        for letter, target in day_column_widths.items():
            current_dim = sheet.column_dimensions.get(letter)
            current_width = getattr(current_dim, "width", None) if current_dim is not None else None
            if current_width is None or current_width < target:
                sheet.column_dimensions[letter].width = target

        off_col = max_col + 2
        off_row = 3

        dec_title = sheet.cell(row=off_row, column=off_col, value="Tableau de dÃ©compte")
        dec_title.fill      = header_fill
        dec_title.alignment = Alignment(horizontal="center")
        dec_title.border    = thin_border
        dec_title.font      = Font(bold=True)

        cols = getattr(shift_count_table, "columns", []) or []
        export_cols = list(cols)
        for ci, nm in enumerate(export_cols):
            ch = sheet.cell(row=off_row+1, column=off_col+ci, value=nm)
            ch.fill      = header_fill
            ch.alignment = Alignment(horizontal="center", vertical="center")
            ch.border    = thin_border
            ch.font      = Font(bold=True)
            letter = get_column_letter(off_col+ci)
            if ci == 0:
                width = 22
            else:
                width = 14
            sheet.column_dimensions[letter].width = width

        tree = getattr(shift_count_table, "tree", None)
        items = tree.get_children() if tree is not None else []

        for i, it in enumerate(items, start=1):
            tags = tree.item(it, 'tags')
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            if "oddrow" in tags:
                fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            vals = tree.item(it, 'values') if tree is not None else []
            row_values = list(vals)
            for ci, v in enumerate(row_values):
                cd = sheet.cell(row=off_row+1+i, column=off_col+ci, value=v)
                cd.alignment = Alignment(horizontal="center", vertical="center")
                cd.border    = thin_border
                cd.fill      = fill
                if ci == 0:
                    cd.font = Font(bold=True)

        # ===========================================================
        #               STATISTIQUES INDIVIDUELLES + PIZZA
        # ===========================================================
        stats_start_row = off_row + 1 + len(items) + 2  # une ligne vide aprÃ¨s le dÃ©compte

        # Titre "Statistiques individuelles"
        sheet.merge_cells(start_row=stats_start_row, start_column=off_col,
                          end_row=stats_start_row,   end_column=off_col+1)
        st_title = sheet.cell(row=stats_start_row, column=off_col, value="Statistiques individuelles")
        st_title.fill      = header_fill
        st_title.alignment = Alignment(horizontal="center", vertical="center")
        st_title.border    = thin_border
        st_title.font      = Font(bold=True)

        # Colonnes confort (tableau stats)
        sheet.column_dimensions[get_column_letter(off_col)].width   = max(24, sheet.column_dimensions.get(get_column_letter(off_col), type('', (), {'width':0})()).width or 0)
        sheet.column_dimensions[get_column_letter(off_col+1)].width = max(14, sheet.column_dimensions.get(get_column_letter(off_col+1), type('', (), {'width':0})()).width or 0)

        # ---- Collecte des personnes (via dÃ©compte + planning) ----
        people_set = set()
        if tree is not None:
            for it in items:
                vals = tree.item(it, 'values')
                if vals:
                    people_set.add(str(vals[0]).strip())
        for di in range(day_count):
            if di >= len(gui_instance.table_entries):
                break
            row_entries = gui_instance.table_entries[di]
            for p_index, _post in enumerate(work_posts):
                if p_index >= len(row_entries):
                    continue
                try:
                    val = row_entries[p_index].get().strip()
                except Exception:
                    val = ""
                if val:
                    people_set.add(val)

        per_person_post_counts = {p: Counter() for p in people_set}
        for di in range(day_count):
            if di >= len(gui_instance.table_entries):
                break
            row_entries = gui_instance.table_entries[di]
            for p_index, post in enumerate(work_posts):
                if p_index >= len(row_entries):
                    continue
                try:
                    val = row_entries[p_index].get().strip()
                except Exception:
                    val = ""
                if val:
                    per_person_post_counts[val][post] += 1

        row_ptr = stats_start_row + 2
        for person in sorted(people_set, key=lambda s: s):  # ordre alpha
            # Titre personne
            sheet.merge_cells(start_row=row_ptr, start_column=off_col,
                              end_row=row_ptr,   end_column=off_col+1)
            name_cell = sheet.cell(row=row_ptr, column=off_col, value=person)
            name_cell.fill      = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
            name_cell.border    = thin_border
            name_cell.font      = Font(bold=True)
            row_ptr += 1

            # Tableau "par poste exact" dans l'ordre de work_posts
            for post in work_posts:
                n = per_person_post_counts.get(person, Counter()).get(post, 0)
                if n > 0:
                    lab = sheet.cell(row=row_ptr, column=off_col,   value=post)
                    val = sheet.cell(row=row_ptr, column=off_col+1, value=f"{n} vacs")
                    for c in (lab, val):
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        c.border    = thin_border
                    row_ptr += 1

            # Moyenne d'astreintes par mois (basée sur le cumul multi-onglets)
            week_vector = weekly_totals_by_person.get(norm_person(person), [0] * n_weeks)
            if n_weeks > 0:
                avg_per_month = round((sum(week_vector) / n_weeks) * 4.0, 1)
            else:
                avg_per_month = 0
            lab = sheet.cell(row=row_ptr, column=off_col,   value="Moyenne d'astreintes / mois")
            val = sheet.cell(row=row_ptr, column=off_col+1, value=avg_per_month)
            for c in (lab, val):
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = thin_border
            row_ptr += 1

            # Ligne vide entre personnes
            row_ptr += 1

        for letter, target in day_column_widths.items():
            current_dim = sheet.column_dimensions.get(letter)
            current_width = getattr(current_dim, "width", None) if current_dim is not None else None
            if current_width is None or current_width < target:
                sheet.column_dimensions[letter].width = target

    # --- Sauvegarde ---
    try:
        workbook.save(file_path)
        messagebox.showinfo("Export", f"Planning exporté dans {file_path}")
        return file_path
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de l'exportation : {e}")
        return None

def export_combined_to_excel(root, tabs_data, days, work_posts, POST_INFO):
    """
    Export 'combinÃ©' : reprend lâ€™export Excel standard (couleurs, largeurs, PDS, Absence),
    mais dans chaque demi-Journée, ajoute en DEUXIÃˆME ligne lâ€™initiale venant dâ€™un autre
    planning .pkl (ex : planning des internes). Ne gÃ©nÃ¨re PAS le tableau de dÃ©compte ni les stats.
    Ajoute en bas une 2e ligne d'absences : 'Absence Planning2' (absences du .pkl importÃ©).
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from tkinter import messagebox, filedialog
    import pickle, os

    # --- Choisir le .pkl Ã  combiner ---
    pkl_path = filedialog.askopenfilename(
        parent=root,
        title="Sélectionner le planning Ã  combiner (.pkl)",
        filetypes=[("Fichiers PKL", "*.pkl"), ("Tous les fichiers", "*.*")]
    )
    if not pkl_path:
        return None

    # --- Choisir le fichier Excel de sortie ---
    file_path = filedialog.asksaveasfilename(
        parent=root,
        title="Exporter le planning (combinÃ©)",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not file_path:
        return None

    # --- Charger le planning importÃ© ---
    try:
        with open(pkl_path, "rb") as f:
            # format : (all_week_status, work_posts, POST_INFO, options)
            # all_week_status[i] : (table_data, cell_av, constraints_data, schedule_data, week_label, excluded)
            other_all_week_status, other_posts, _other_post_info, _other_opts = pickle.load(f)
    except Exception as e:
        messagebox.showerror("Export combiné", f"Impossible de lire le .pkl sélectionnée :\n{e}")
        return None

    # -------- Helpers --------
    def _norm(s: str) -> str:
        return " ".join(str(s or "").split()).lower()

    # Mapping des semaines (par libellÃ© du label de semaine ; sinon fallback par index)
    cur_week_labels = []
    for (gui_w, _c, _t) in tabs_data:
        try:
            cur_week_labels.append(gui_w.week_label.cget("text").strip())
        except Exception:
            cur_week_labels.append("")

    other_week_labels = []
    for wk in other_all_week_status:
        lab = ""
        try:
            lab = wk[4] if len(wk) >= 5 else ""  # 5e champ = texte du label si prÃ©sent
        except Exception:
            pass
        other_week_labels.append((lab or "").strip())

    other_index_by_label = {_norm(l): i for i, l in enumerate(other_week_labels)}
    week_map = []
    for i, lab in enumerate(cur_week_labels):
        j = other_index_by_label.get(_norm(lab))
        if j is None:
            j = i if i < len(other_all_week_status) else None
        week_map.append(j)

    # Mapping des postes (par nom exact, sinon fallback index)
    other_post_index = {p: i for i, p in enumerate(other_posts)}
    day_count = len(days)

    # -------- Classeur & styles (alignÃ©s sur lâ€™export standard) --------
    workbook = openpyxl.Workbook()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    dark_grey_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")

    # Nom simple pour le 2e planning (affichage info)
    planning2_name = os.path.basename(pkl_path)

    for idx, (gui_instance, constraints_instance, _shift_table) in enumerate(tabs_data):
        # 1 feuille par semaine
        sheet = workbook.active if idx == 0 else workbook.create_sheet(title=f"Semaine {idx+1}")
        if idx == 0:
            sheet.title = f"Semaine {idx+1}"

        day_column_widths = {}

        # ---------- Collecte PDS (planning courant) ----------
        pds_by_day = [[] for _ in range(day_count)]
        for row in getattr(constraints_instance, "rows", []):
            if not row or len(row) < 5:
                continue
            try:
                initials = row[0].get().strip()
            except Exception:
                initials = ""
            max_di = min(day_count, len(row) - 4)
            for di in range(max_di):
                tpl = row[4 + di]
                if isinstance(tpl, tuple) and len(tpl) == 3:
                    try:
                        if tpl[2].get() == 1:
                            pds_by_day[di].append(initials)
                    except Exception:
                        pass

        # Titre semaine (A1 simple)
        try:
            semaine_label = gui_instance.week_label.cget("text").strip()
        except Exception:
            semaine_label = f"Semaine {idx+1}"
        sheet.cell(row=1, column=1, value=semaine_label).border = thin_border
        sheet.cell(row=1, column=3, value=f"CombinÃ© avec : {planning2_name}").border = thin_border

        col_post_name = 1
        col_shift     = 2
        col_day_start = 3
        header_row    = 2

        # En-tÃªtes jours
        for di in range(day_count):
            x = col_day_start + di
            c = sheet.cell(row=header_row, column=x, value=days[di])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border
            c.font      = Font(bold=True)

        row_offset = 3

        # Table & contraintes du planning importÃ© pour cette semaine (si dispo)
        other_idx = week_map[idx]
        other_table = None
        other_constraints = None
        if other_idx is not None and other_idx < len(other_all_week_status):
            try:
                other_table = other_all_week_status[other_idx][0]      # table_data (2*posts x 7)
            except Exception:
                other_table = None
            try:
                other_constraints = other_all_week_status[other_idx][2]  # constraints_data (rows)
            except Exception:
                other_constraints = None

        # ---------- Corps planning (matin + aprÃ¨s-midi) ----------
        for p_index, post in enumerate(work_posts):
            hexcol = (POST_INFO.get(post, {}).get("color", "#DDDDDD") or "#DDDDDD").lstrip("#")
            post_fill = PatternFill(start_color=hexcol, end_color=hexcol, fill_type="solid")

            # Nom du poste (fusion 4 lignes)
            sheet.merge_cells(start_row=row_offset,   start_column=col_post_name,
                              end_row=row_offset+3,   end_column=col_post_name)
            cp = sheet.cell(row=row_offset, column=col_post_name, value=post)
            cp.fill = post_fill; cp.alignment = Alignment(horizontal="center", vertical="center")
            cp.border = thin_border; cp.font = Font(bold=True)

            # LibellÃ©s "MATIN" / "AP MIDI"
            sheet.merge_cells(start_row=row_offset,   start_column=col_shift,
                              end_row=row_offset+1,   end_column=col_shift)
            lm = sheet.cell(row=row_offset, column=col_shift, value="MATIN")
            lm.fill = post_fill; lm.alignment = Alignment(horizontal="center", vertical="center")
            lm.border = thin_border; lm.font = Font(bold=True)

            sheet.merge_cells(start_row=row_offset+2, start_column=col_shift,
                              end_row=row_offset+3,   end_column=col_shift)
            la = sheet.cell(row=row_offset+2, column=col_shift, value="AP MIDI")
            la.fill = post_fill; la.alignment = Alignment(horizontal="center", vertical="center")
            la.border = thin_border; la.font = Font(bold=True)

            # Lignes du planning courant
            rg_matin, rg_apmidi = p_index*2, p_index*2 + 1

            # Index poste pour le planning importÃ©
            other_p_index = other_post_index.get(post)
            if other_p_index is None and p_index < len(other_posts):
                other_p_index = p_index  # fallback si noms diffÃ©rents

            for di in range(day_count):
                x = col_day_start + di

                # ----- MATIN -----
                lbl_m = gui_instance.table_labels[rg_matin][di]
                ent_m = gui_instance.table_entries[rg_matin][di]
                txt_time  = lbl_m.cget("text") if lbl_m else ""
                txt_init  = ent_m.get().strip() if ent_m else ""
                other_init_m = ""
                is_disabled_m = not gui_instance.cell_availability.get((rg_matin, di), True)
                if other_table is not None and other_p_index is not None:
                    try:
                        other_init_m = str(other_table[other_p_index*2][di] or "").strip()
                    except Exception:
                        pass

                has_person_m = bool(txt_init or other_init_m)
                slot_open_m = (not is_disabled_m) or has_person_m
                ctime = sheet.cell(row=row_offset, column=x, value=txt_time)
                register_day_column_width(day_column_widths, x, txt_time)
                if slot_open_m:
                    ctime.fill = post_fill;      ctime.font = Font(color="000000")
                else:
                    ctime.fill = dark_grey_fill; ctime.font = Font(color="999999")
                ctime.alignment = Alignment(horizontal="center", vertical="center")
                ctime.border    = thin_border

                combined = txt_init + (("\n" + other_init_m) if other_init_m else "")
                cinits = sheet.cell(row=row_offset+1, column=x, value=combined)
                register_day_column_width(day_column_widths, x, combined)
                if slot_open_m:
                    cinits.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    if combined:
                        cinits.font = Font(bold=True, color="FF0000")
                else:
                    cinits.fill = dark_grey_fill; cinits.font = Font(color="000000")
                cinits.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cinits.border    = thin_border
                sheet.row_dimensions[row_offset+1].height = 28  # 2 lignes

                # ----- AP MIDI -----
                lbl_a = gui_instance.table_labels[rg_apmidi][di]
                ent_a = gui_instance.table_entries[rg_apmidi][di]
                txt_time2 = lbl_a.cget("text") if lbl_a else ""
                txt_init2 = ent_a.get().strip() if ent_a else ""
                other_init_a = ""
                if other_table is not None and other_p_index is not None:
                    try:
                        other_init_a = str(other_table[other_p_index*2 + 1][di] or "").strip()
                    except Exception:
                        pass
                is_disabled_a = not gui_instance.cell_availability.get((rg_apmidi, di), True)

                has_person_a = bool(txt_init2 or other_init_a)
                slot_open_a = (not is_disabled_a) or has_person_a
                ctime2 = sheet.cell(row=row_offset+2, column=x, value=txt_time2)
                register_day_column_width(day_column_widths, x, txt_time2)
                if slot_open_a:
                    ctime2.fill = post_fill;      ctime2.font = Font(color="000000")
                else:
                    ctime2.fill = dark_grey_fill; ctime2.font = Font(color="999999")
                ctime2.alignment = Alignment(horizontal="center", vertical="center")
                ctime2.border    = thin_border

                combined2 = txt_init2 + (("\n" + other_init_a) if other_init_a else "")
                cinits2 = sheet.cell(row=row_offset+3, column=x, value=combined2)
                register_day_column_width(day_column_widths, x, combined2)
                if slot_open_a:
                    cinits2.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    if combined2:
                        cinits2.font = Font(bold=True, color="FF0000")
                else:
                    cinits2.fill = dark_grey_fill; cinits2.font = Font(color="000000")
                cinits2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cinits2.border    = thin_border
                sheet.row_dimensions[row_offset+3].height = 28

            row_offset += 4

        # Largeurs de colonnes
        max_col = col_day_start + day_count - 1
        for c in range(1, max_col+1):
            let = get_column_letter(c)
            sheet.column_dimensions[let].width = 20 if c == 1 else 15

        # ---------- Lignes PDS / Absence (planning courant) ----------
        pds_row = row_offset + 1
        pds_title = sheet.cell(row=pds_row, column=col_shift, value="PDS")
        pds_title.font = Font(bold=True)
        pds_title.alignment = Alignment(horizontal="center", vertical="center")
        pds_title.border = thin_border
        for di in range(day_count):
            x = col_day_start + di
            val = ", ".join(pds_by_day[di]) if pds_by_day[di] else ""
            c = sheet.cell(row=pds_row, column=x, value=val)
            register_day_column_width(day_column_widths, x, val)
            c.font      = Font(bold=True, color="FF0000") if val else Font(color="000000")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border
            c.fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        abs_row = pds_row + 1
        abs_title = sheet.cell(row=abs_row, column=col_shift, value="Absence")
        abs_title.font = Font(bold=True)
        abs_title.alignment = Alignment(horizontal="center", vertical="center")
        abs_title.border = thin_border

        abs_by_day = [[] for _ in range(day_count)]
        for row in getattr(constraints_instance, "rows", []):
            if not row or len(row) < 5:
                continue
            try:
                ini = row[0].get().strip()
            except Exception:
                ini = ""
            max_di = min(day_count, len(row) - 4)
            for di in range(max_di):
                tpl = row[4 + di]
                if isinstance(tpl, tuple) and tpl:
                    try:
                        st = tpl[0]._var.get().strip().upper()
                    except Exception:
                        st = ""
                    if st:
                        abs_by_day[di].append(f"{ini} ({st})")

        for di in range(day_count):
            x = col_day_start + di
            s = "\n".join(abs_by_day[di]) if abs_by_day[di] else ""
            c = sheet.cell(row=abs_row, column=x, value=s)
            register_day_column_width(day_column_widths, x, s)
            c.font      = Font(bold=True, color="FF0000") if s else Font(color="000000")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = thin_border
            c.fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ---------- NOUVEAU : 'Absence Planning2' (planning importÃ©) ----------
        abs2_row = abs_row + 1
        abs2_title = sheet.cell(row=abs2_row, column=col_shift, value="Absence Planning2")
        abs2_title.font = Font(bold=True)
        abs2_title.alignment = Alignment(horizontal="center", vertical="center")
        abs2_title.border = thin_border

        other_abs_by_day = [[] for _ in range(day_count)]
        if other_constraints and isinstance(other_constraints, list):
            for row_data in other_constraints:
                try:
                    ini2 = (row_data[0] or "").strip()
                except Exception:
                    ini2 = ""
                # colonnes jours 4..(4+day_count-1) : tuples (absence, pds)
                for di in range(day_count):
                    col = 4 + di
                    if col < len(row_data):
                        cell = row_data[col]
                        if isinstance(cell, (list, tuple)) and cell:
                            try:
                                st2 = (cell[0] or "").strip().upper()
                            except Exception:
                                st2 = ""
                            if st2:
                                other_abs_by_day[di].append(f"{ini2} ({st2})")

        for di in range(day_count):
            x = col_day_start + di
            s2 = "\n".join(other_abs_by_day[di]) if other_abs_by_day[di] else ""
            c2 = sheet.cell(row=abs2_row, column=x, value=s2)
            register_day_column_width(day_column_widths, x, s2)
            c2.font      = Font(bold=True, color="FF0000") if s2 else Font(color="000000")
            c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c2.border    = thin_border
            c2.fill      = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # >>> PAS de tableau de dÃ©compte, PAS de statistiques individuelles <<<

    # --- Sauvegarde ---
    try:
        workbook.save(file_path)
        base = os.path.basename(pkl_path)
        messagebox.showinfo("Export combiné", f"Planning exporté (combiné avec '{base}') dans :\n{file_path}")
        return file_path
    except Exception as e:
        messagebox.showerror("Erreur", f"Erreur lors de l'export combiné : {e}")
        return None

